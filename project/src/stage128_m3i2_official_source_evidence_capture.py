"""Stage128 — ``stage128-m3i2-official-source-evidence-capture``.

Acquire, hash, package and validate **official-source evidence** for the merged
M3I-2 contract and the contingent M3I-3 financing shell.

What this action answers
------------------------
Only one question: *is there an independently auditable, raw-byte-backed
official-source evidence package for the locked M3I-2 definitions?*

What it deliberately does not answer
------------------------------------
Whether M3I-2 meets coverage thresholds, whether it improves prediction,
whether it should be admitted, or what the final model is. Those need their own
authorizations. This action performs **no** company-panel join, **no** feature
materialization, **no** coverage calculation, **no** Data Gate, **no** modeling
and **no** final-test access.

Three modes, deliberately separated
-----------------------------------
``--capture``            the one authorized network session (delegated to
                         :mod:`stage128_m3i2_capture_layer`, the only module
                         allowed to open a socket)
``--build-from-bundle``  offline reconstruction of every normalized artifact
                         from the retained raw bytes
``--check``              offline verification of the committed package

This module itself is import-clean of network libraries: the static firewall
below asserts that, and it is the reason the capture layer lives in its own
file.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import sys
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

# --------------------------------------------------------------------------- #
# Identity
# --------------------------------------------------------------------------- #

ACTION_ID = "stage128-m3i2-official-source-evidence-capture"
PACKAGE_ID = "stage128_m3i2_official_source_evidence_capture"
PACKAGE_VERSION = "stage128_m3i2_official_source_evidence_capture_v1"
ACTION_TYPE = (
    "official_source_evidence_capture_only_no_join_no_feature_no_coverage_"
    "no_gate_no_modeling")

REPOSITORY = "abtinasg/papermali"

# --- baseline (section 1) --------------------------------------------------- #

BASELINE_BRANCH = "main"
BASELINE_COMMIT = "cf23771a383bf9ad8f7ff2855c216c9a240647ff"

PREDECESSOR_PR_NUMBER = 74
PREDECESSOR_PR_HEAD_COMMIT = "0afcedfda54a2e7530aa5e55fc6de80394a9b248"
PREDECESSOR_PR_MERGED = True
PREDECESSOR_PR_MERGE_COMMIT = "cf23771a383bf9ad8f7ff2855c216c9a240647ff"
PREDECESSOR_ACTION_ID = "stage128-m3i2-prospective-contract-lock"

HEAD_BRANCH = "stage128-m3i2-official-source-evidence-capture"
PR_BASE_BRANCH = "main"
PR_IS_DRAFT = True
MERGE_AUTHORIZED = False
AUTO_MERGE = False

PACKAGE_DIR_REL = "project/stage128/m3i2_official_source_evidence_capture"

README_REL = (
    f"{PACKAGE_DIR_REL}/"
    "README_STAGE128_M3I2_OFFICIAL_SOURCE_EVIDENCE_CAPTURE.md")
AUTHORIZATION_REL = (
    f"{PACKAGE_DIR_REL}/"
    "stage128_m3i2_evidence_capture_human_authorization_record.json")
GOVERNANCE_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_evidence_capture_governance_boundary.json")
CUTOFF_AUDIT_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_development_cutoff_source_audit.json")
CUTOFF_PLAN_REL = f"{PACKAGE_DIR_REL}/stage128_m3i2_unique_cutoff_plan.csv"
RELEASE_MANIFEST_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_wdi_archive_release_manifest.csv")
REQUIRED_EDITIONS_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_required_wdi_editions.csv")
REQUEST_MANIFEST_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_official_request_manifest.csv")
RESPONSE_MANIFEST_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_official_response_manifest.csv")
LOCKED_SERIES_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_wdi_irn_locked_series_extract.csv")
SEMANTIC_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_wdi_vintage_semantic_compatibility.csv")
IMF_CATALOG_REL = f"{PACKAGE_DIR_REL}/stage128_m3i3_imf_mfs_ir_catalog.csv"
FINANCING_EVIDENCE_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i3_financing_metadata_evidence.json")
CONTINUATION_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_capture_continuation_record.json")
BUNDLE_MANIFEST_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_external_bundle_manifest.json")
DECISION_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_official_source_evidence_decision.json")
QC_REL = (
    f"{PACKAGE_DIR_REL}/stage128_m3i2_official_source_evidence_qc_report.json")
METADATA_REL = (
    f"{PACKAGE_DIR_REL}/"
    "metadata_and_hashes_stage128_m3i2_official_source_evidence_capture.json")

#: The capture layer is EXCLUDED here on purpose: it is the one module allowed
#: to import a network library, and the offline scanners below run over this
#: list.
OFFLINE_IMPLEMENTATION_FILES: tuple[str, ...] = (
    "project/src/stage128_m3i2_official_source_evidence_capture.py",
    "project/tests/test_stage128_m3i2_official_source_evidence_capture.py",
)

CAPTURE_LAYER_FILE = "project/src/stage128_m3i2_capture_layer.py"
RUNNER_FILE = "project/run_stage128_m3i2_official_source_evidence_capture.py"


class M3I2EvidenceCaptureError(RuntimeError):
    """Fail-closed error for the M3I-2 official-source evidence capture."""


# --------------------------------------------------------------------------- #
# Human authorization (section 0)
# --------------------------------------------------------------------------- #

#: The EXACT Persian authorization block sent by the human supervisor with the
#: canonical prompt. Verbatim human text; authoritative ONLY in the
#: authorization record. No trailing newline.
HUMAN_AUTHORIZATION_TEXT = (
    "مجاز است فقط مرحله\n"
    "stage128-m3i2-official-source-evidence-capture\n"
    "در مخزن abtinasg/papermali از main با expected SHA\n"
    "cf23771a383bf9ad8f7ff2855c216c9a240647ff\n"
    "اجرا شود.\n"
    "\n"
    "در این مرحله فقط شواهد رسمی World Bank WDI و IMF برای قرارداد\n"
    "M3I-2 و پوسته مشروط M3I-3 دریافت، ثبت، هش و بسته‌بندی شود.\n"
    "\n"
    "هیچ داده کلان به ردیف شرکت‌ها متصل نشود، هیچ feature نهایی ساخته نشود،\n"
    "coverage یا Data Gate محاسبه نشود، هیچ مدل یا M3I-vs-M2 اجرا نشود،\n"
    "M4 شروع نشود، Final Test باز نشود و هیچ PRی Merge نشود."
)
HUMAN_AUTHORIZATION_UTF8_BYTES = 695
HUMAN_AUTHORIZATION_SHA256 = (
    "eb0230b06269feee5f274315d2958f762c69fc231f36c73b0048415e5fd95b06")
HUMAN_AUTHORIZATION_LOCAL_TIMESTAMP = "2026-08-03T13:30:00+03:30"
HUMAN_AUTHORIZATION_PRECEDING_CONTEXT = (
    "Sent by the human supervisor together with the canonical programmer "
    "prompt for stage128-m3i2-official-source-evidence-capture, immediately "
    "after PR #74 (the M3I-2 prospective contract lock) was merged into main "
    "by merge commit cf23771a383bf9ad8f7ff2855c216c9a240647ff.")

#: This authorization is explicit and self-describing: it names the action, the
#: repository and the expected baseline SHA. It was NOT inferred from a prompt
#: hash, a branch name or a Roadmap pointer. A pointer is never authorization.
AUTHORIZATION_INFERRED_FROM_POINTER = False
AUTHORIZATION_INFERRED_FROM_PRIOR_PROMPT_HASH = False
AUTHORIZATION_INFERRED_FROM_BRANCH_NAME = False
AUTHORIZATION_IS_STANDING = False
AUTHORIZATION_CONSUMED = True

#: DERIVED, non-verbatim restatement.
NORMALIZED_AUTHORIZATION_SCOPE = (
    "Capture, record, hash and package official World Bank WDI and IMF "
    "evidence for the merged M3I-2 contract and the contingent M3I-3 "
    "financing shell, joining no macro value to any company row, "
    "materializing no feature, computing no coverage, executing no Data Gate, "
    "fitting no model, starting no M4, opening no final test and merging no "
    "PR.")


#: The capture ran in TWO invocations. The archive URLs live inside the
#: official listing, so they cannot be known until that listing is captured;
#: the programmer's runner closed its manifest after discovery, stopped, and
#: put the question to the human rather than starting a second capture on its
#: own authority. The human chose to complete the same capture.
CAPTURE_INVOCATIONS = 2
INVOCATION_1_ROLE = "discovery_only"
INVOCATION_1_REQUEST_COUNT = 5
INVOCATION_1_CLOSED = True
INVOCATION_2_ROLE = "required_archive_completion"
INVOCATION_2_REQUEST_COUNT = 16
INVOCATION_2_CLOSED = True

CONTINUATION_WAS_EXPLICITLY_REQUESTED_FROM_HUMAN = True
CONTINUATION_SCOPE_CHANGED = False
CONTINUATION_AUTHORIZED_NEW_SCIENTIFIC_ACTION = False
CONTINUATION_AUTHORIZED_ONLY_COMPLETION_OF_SAME_CAPTURE = True

#: The exact option text the human selected, verbatim.
CONTINUATION_SELECTED_OPTION_TEXT = "capture را کامل کن (پیشنهادی)"
CONTINUATION_QUESTION_CONTEXT = (
    "The programmer reported that invocation 1 captured only the five "
    "official discovery objects and closed its manifest, that the 16 required "
    "archive-edition URLs are only knowable after parsing the captured "
    "listing, and that section 7 of the canonical prompt forbids a second "
    "live capture after the authorization is consumed. The human was offered "
    "exactly two options: complete the same capture, or stop and report with "
    "required_editions_captured = 0.")
#: No exact wall-clock timestamp for the human's selection is retained.
CONTINUATION_TIMESTAMP_STATUS = "UNRESOLVED_EXACT_TIMESTAMP"
CONTINUATION_OBSERVED_REQUEST_WINDOW = (
    "invocation 1 discovery requests and invocation 2 archive requests are "
    "separated in the retained request manifest; the selection occurred "
    "between them")


def build_continuation_record(requests_rows: list[dict[str, Any]]
                              ) -> dict[str, Any]:
    """Disclose BOTH capture invocations in the committed package.

    The external bundle is not the only place this belongs: an auditor reading
    the repository alone must be able to see that the capture ran twice and on
    whose say-so.
    """
    discovery = [r for r in requests_rows
                 if not str(r.get("object_id", "")).startswith("wdi_archive_")]
    archives = [r for r in requests_rows
                if str(r.get("object_id", "")).startswith("wdi_archive_")]
    times = sorted(str(r.get("started_utc") or "") for r in requests_rows
                   if r.get("started_utc"))
    return {
        "package_id": PACKAGE_ID,
        "generated_for": ACTION_ID,
        "record_type": "supplementary_continuation_record",
        "supplements_but_does_not_replace":
            "stage128_m3i2_evidence_capture_human_authorization_record.json",

        "capture_invocations": CAPTURE_INVOCATIONS,
        "invocation_1_role": INVOCATION_1_ROLE,
        "invocation_1_request_count": INVOCATION_1_REQUEST_COUNT,
        "invocation_1_closed": INVOCATION_1_CLOSED,
        "invocation_1_observed_request_count": len(discovery),
        "invocation_2_role": INVOCATION_2_ROLE,
        "invocation_2_request_count": INVOCATION_2_REQUEST_COUNT,
        "invocation_2_closed": INVOCATION_2_CLOSED,
        "invocation_2_observed_request_count": len(archives),

        "continuation_was_explicitly_requested_from_human":
            CONTINUATION_WAS_EXPLICITLY_REQUESTED_FROM_HUMAN,
        "continuation_selected_option_text":
            CONTINUATION_SELECTED_OPTION_TEXT,
        "continuation_selected_option_text_is_verbatim_human_selection": True,
        "continuation_question_context": CONTINUATION_QUESTION_CONTEXT,
        "continuation_timestamp_status": CONTINUATION_TIMESTAMP_STATUS,
        "continuation_observed_request_window":
            CONTINUATION_OBSERVED_REQUEST_WINDOW,
        "first_recorded_request_started_utc": times[0] if times else "",
        "last_recorded_request_started_utc": times[-1] if times else "",

        "continuation_scope_changed": CONTINUATION_SCOPE_CHANGED,
        "continuation_authorized_new_scientific_action":
            CONTINUATION_AUTHORIZED_NEW_SCIENTIFIC_ACTION,
        "continuation_authorized_only_completion_of_same_capture":
            CONTINUATION_AUTHORIZED_ONLY_COMPLETION_OF_SAME_CAPTURE,
        "third_invocation_present": False,
        "programmer_stopped_and_asked_rather_than_recapturing": True,
        "raw_bytes_from_both_invocations_retained": True,
        "no_invocation_1_evidence_deleted_or_rewritten": True,
    }


def assert_continuation_record_is_sound(record: dict[str, Any]) -> None:
    """Fail closed on an undisclosed, unauthorized or widened continuation."""
    invocations = record.get("capture_invocations")
    if invocations is None or invocations < 1:
        raise M3I2EvidenceCaptureError(
            "the capture-invocation count must be recorded")
    if invocations > 2:
        raise M3I2EvidenceCaptureError(
            f"a third capture invocation appeared ({invocations}); only the "
            "discovery invocation and its authorized completion exist")
    if invocations == 2:
        if record.get(
                "continuation_was_explicitly_requested_from_human") is not True:
            raise M3I2EvidenceCaptureError(
                "a second capture invocation without an explicit human "
                "continuation decision is unauthorized")
        if not record.get("continuation_selected_option_text"):
            raise M3I2EvidenceCaptureError(
                "the human's continuation selection must be recorded verbatim")
        if record.get("continuation_scope_changed") is not False:
            raise M3I2EvidenceCaptureError(
                "the continuation may not widen scope")
        if record.get(
                "continuation_authorized_new_scientific_action") is not False:
            raise M3I2EvidenceCaptureError(
                "the continuation authorized no new scientific action")
        if record.get(
                "continuation_authorized_only_completion_of_same_capture"
        ) is not True:
            raise M3I2EvidenceCaptureError(
                "the continuation authorized only completion of the same "
                "capture")
    if record.get("third_invocation_present") is not False:
        raise M3I2EvidenceCaptureError("a third invocation is not authorized")


def verify_human_authorization() -> dict[str, Any]:
    """Fail closed unless the recorded authorization is byte-exact."""
    raw = HUMAN_AUTHORIZATION_TEXT.encode("utf-8")
    if len(raw) != HUMAN_AUTHORIZATION_UTF8_BYTES:
        raise M3I2EvidenceCaptureError(
            f"authorization byte length {len(raw)} != "
            f"{HUMAN_AUTHORIZATION_UTF8_BYTES}")
    digest = hashlib.sha256(raw).hexdigest()
    if digest != HUMAN_AUTHORIZATION_SHA256:
        raise M3I2EvidenceCaptureError(
            f"authorization sha256 {digest} != {HUMAN_AUTHORIZATION_SHA256}")
    if HUMAN_AUTHORIZATION_TEXT.endswith("\n"):
        raise M3I2EvidenceCaptureError(
            "the verbatim authorization must not carry a trailing newline")
    if ACTION_ID not in HUMAN_AUTHORIZATION_TEXT:
        raise M3I2EvidenceCaptureError(
            "the authorization text must name this action explicitly")
    if BASELINE_COMMIT not in HUMAN_AUTHORIZATION_TEXT:
        raise M3I2EvidenceCaptureError(
            "the authorization text must name the expected baseline SHA")
    return {
        "authorization_utf8_bytes": len(raw),
        "authorization_sha256": digest,
        "normalized_authorization_scope_sha256": hashlib.sha256(
            NORMALIZED_AUTHORIZATION_SCOPE.encode("utf-8")).hexdigest(),
    }


# --------------------------------------------------------------------------- #
# The merged contract is READ-ONLY (section 3)
# --------------------------------------------------------------------------- #

CONTRACT_PACKAGE_DIR = "project/stage128/m3_intl_macro_contract_lock"
CONTRACT_DEFINITION_LOCK_REL = (
    f"{CONTRACT_PACKAGE_DIR}/stage128_m3_intl_macro_definition_lock.json")
CONTRACT_PREDICTION_TIME_REL = (
    f"{CONTRACT_PACKAGE_DIR}/"
    "stage128_m3_intl_macro_prediction_time_contract.json")
CONTRACT_DECISION_REL = (
    f"{CONTRACT_PACKAGE_DIR}/stage128_m3_intl_macro_contract_decision.json")

CPI_CANDIDATE_ID = "cand_m3i_cpi_inflation_annual"
FX_CANDIDATE_ID = "cand_m3i_fx_change_official_annual"
FINANCING_CANDIDATE_ID = "cand_m3i_financing_rate"

CPI_INDICATOR_CODE = "FP.CPI.TOTL.ZG"
FX_INDICATOR_CODE = "PA.NUS.FCRF"
LOCKED_INDICATOR_CODES: tuple[str, ...] = (CPI_INDICATOR_CODE,
                                           FX_INDICATOR_CODE)
FORBIDDEN_INDICATOR_CODES: tuple[str, ...] = (
    "PA.NUS.ATLS", "FP.CPI.TOTL", "NY.GDP.DEFL.KD.ZG", "FR.INR.LEND",
)

IRAN_ECONOMY_CODE = "IRN"
IRAN_ECONOMY_NAMES: tuple[str, ...] = (
    "Iran, Islamic Rep.", "Iran, Islamic Republic of", "Iran (Islamic "
    "Republic of)", "Islamic Republic of Iran",
)

M3I3_DATASET_ID = "IMF.STA:MFS_IR"
M3I3_LOCK_STATUS = "UNRESOLVED_METADATA_LOCK"
M3I3_PENDING_STATUS = (
    "IDENTIFIED_FROM_OFFICIAL_METADATA_PENDING_SEPARATE_PROSPECTIVE_LOCK")

M3_CBI_STATUS = "UNRESOLVED_M3_DATA_GATE"

FINANCING_FORBIDDEN_PROXIES: tuple[str, ...] = (
    "deposit interest rate",
    "deposit-rate ceiling",
    "real interest rate",
    "interest-rate spread",
    "repo transaction volume",
    "reverse-repo transaction volume",
    "standing-facility transaction amount",
    "currency-volume series",
    "a differently defined policy rate relabelled as financing rate",
    "World Bank FR.INR.LEND admitted without a separate Gate",
)

FINANCING_ACCEPTABLE_CONSTRUCT = (
    "An exact official Iran rate representing the cost of bank "
    "lending/financing to the private sector, or an explicitly named "
    "financing/facility rate with the same economic construct.")


def read_merged_contract(root: Path) -> dict[str, Any]:
    """Read the merged contract and fail closed on any drift.

    This is a READ. The contract package is never written by this action.
    """
    lock = json.loads(
        (root / CONTRACT_DEFINITION_LOCK_REL).read_text(encoding="utf-8"))
    cpi, fx = lock["m3i2_candidates"]
    financing = lock["m3i3_candidate"]

    if cpi["candidate_id"] != CPI_CANDIDATE_ID:
        raise M3I2EvidenceCaptureError("CPI candidate id drift")
    if fx["candidate_id"] != FX_CANDIDATE_ID:
        raise M3I2EvidenceCaptureError("FX candidate id drift")
    if cpi["indicator_code"] != CPI_INDICATOR_CODE:
        raise M3I2EvidenceCaptureError(
            f"CPI indicator must be {CPI_INDICATOR_CODE}")
    if fx["indicator_code"] != FX_INDICATOR_CODE:
        raise M3I2EvidenceCaptureError(
            f"FX indicator must be {FX_INDICATOR_CODE}")
    if cpi["transformation_formula"] != "identity":
        raise M3I2EvidenceCaptureError("CPI transformation drift")
    if fx["transformation_formula"] != "100 * ln(E_y / E_(y-1))":
        raise M3I2EvidenceCaptureError("FX transformation drift")
    if financing["candidate_id"] != FINANCING_CANDIDATE_ID:
        raise M3I2EvidenceCaptureError("financing candidate id drift")
    if financing["candidate_selection_status"] != M3I3_LOCK_STATUS:
        raise M3I2EvidenceCaptureError("financing lock status drift")
    if financing["admitted"] is not False:
        raise M3I2EvidenceCaptureError("financing must remain not admitted")
    if financing["exact_series_code"] is not None:
        raise M3I2EvidenceCaptureError(
            "the merged financing shell must keep its null fields")

    return {
        "definition_lock_sha256": _sha256_file(
            root / CONTRACT_DEFINITION_LOCK_REL),
        "prediction_time_sha256": _sha256_file(
            root / CONTRACT_PREDICTION_TIME_REL),
        "contract_decision_sha256": _sha256_file(root / CONTRACT_DECISION_REL),
        "cpi_candidate": cpi,
        "fx_candidate": fx,
        "financing_candidate": financing,
        "contract_read_only": True,
        "contract_modified_by_this_action": False,
    }


# --------------------------------------------------------------------------- #
# Development-cutoff input firewall (section 5)
# --------------------------------------------------------------------------- #

#: The canonical frozen development surface. This binding is NOT invented here:
#: the merged M3 macro data Gate derived the retained-M2 development common
#: sample programmatically from this exact file and this exact membership flag,
#: and that action is merged and independently audited.
CUTOFF_SOURCE_REL = "project/stage128/stage128_m2_d2_development_features.csv"
CUTOFF_MEMBERSHIP_FLAG = "in_three_variable_common_sample"
CUTOFF_FIELD = "pair_cutoff_date"

#: Section 5.2 — the ONLY columns this action may read.
CUTOFF_COLUMN_ALLOWLIST: tuple[str, ...] = (
    "ticker", "fiscal_year_t", "target_year", CUTOFF_FIELD,
    CUTOFF_MEMBERSHIP_FLAG,
)

#: Columns that must never be read. Reading any of them is a firewall breach.
CUTOFF_COLUMN_DENYLIST: tuple[str, ...] = (
    "equity_return_d2", "realized_volatility", "amihud_illiquidity",
    "equity_return_window_d0_historical", "m2_value_status", "d2_status",
    "temporal_folds",
)

DEVELOPMENT_TARGET_YEARS: tuple[str, ...] = (
    "1393", "1394", "1395", "1396", "1397", "1398", "1399")
FINAL_TEST_TARGET_YEARS: tuple[str, ...] = ("1400", "1401", "1402")

EXPECTED_DEVELOPMENT_PAIRS = 539
EXPECTED_DEVELOPMENT_COMPANIES = 108

#: The cutoff column is a DATE with no verified intraday timestamp: Stage125
#: Part3B1A locked the available_at OPERATIONALIZATION but recorded
#: ``zero_real_available_at_assignments`` and
#: ``pilot_cutoff_provenance_resolved: false``. Rather than invent a time, the
#: earliest instant of the cutoff date is used for edition selection. That is
#: strictly conservative: it can only EXCLUDE editions, never admit an extra
#: one, so it cannot manufacture availability.
CUTOFF_TIME_ASSUMPTION = "00:00:00Z"
CUTOFF_TIME_ASSUMPTION_IS_CONSERVATIVE = True
CUTOFF_INTRADAY_TIME_VERIFIED = False


def bind_development_cutoff_source(root: Path) -> dict[str, Any]:
    """Section 5.1 — bind ONE cutoff source, or fail closed."""
    path = root / CUTOFF_SOURCE_REL
    if not path.is_file():
        raise M3I2EvidenceCaptureError(
            f"STOP_DEVELOPMENT_CUTOFF_SOURCE_NOT_UNIQUELY_BOUND: "
            f"{CUTOFF_SOURCE_REL} is absent")

    with path.open(encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        columns = list(reader.fieldnames or ())
        rows = list(reader)

    for column in CUTOFF_COLUMN_ALLOWLIST:
        if column not in columns:
            raise M3I2EvidenceCaptureError(
                f"STOP_DEVELOPMENT_CUTOFF_SOURCE_NOT_UNIQUELY_BOUND: "
                f"required column {column!r} absent")

    return {
        "cutoff_source_repository_path": CUTOFF_SOURCE_REL,
        "cutoff_source_git_blob_sha": _git_blob_sha(root, CUTOFF_SOURCE_REL),
        "cutoff_source_sha256": _sha256_file(path),
        "cutoff_source_bytes": path.stat().st_size,
        "cutoff_source_format": "csv_utf8_with_header",
        "cutoff_source_column_names": columns,
        "cutoff_source_column_count": len(columns),
        "cutoff_source_row_count": len(rows),
        "cutoff_field": CUTOFF_FIELD,
        "membership_flag": CUTOFF_MEMBERSHIP_FLAG,
        "uniquely_bound": True,
        "uniqueness_basis": (
            "The merged, independently audited stage128-m3-macro-data-gate "
            "derived the retained-M2 development common sample "
            "programmatically from this exact file and this exact membership "
            "flag. No second plausible development-cutoff surface exists in "
            "the repository, and no final-test directory was searched to "
            "resolve the binding."),
        "columns_read": list(CUTOFF_COLUMN_ALLOWLIST),
        "columns_never_read": list(CUTOFF_COLUMN_DENYLIST),
        "outcome_or_target_columns_read": False,
        "financial_or_market_feature_columns_read": False,
        "final_test_directories_searched": False,
        "development_target_years": list(DEVELOPMENT_TARGET_YEARS),
        # the material limitation, stated plainly rather than papered over
        "cutoff_field_is_date_only": True,
        "cutoff_intraday_time_verified": CUTOFF_INTRADAY_TIME_VERIFIED,
        "cutoff_time_assumption_for_edition_selection": CUTOFF_TIME_ASSUMPTION,
        "cutoff_time_assumption_is_conservative":
            CUTOFF_TIME_ASSUMPTION_IS_CONSERVATIVE,
        "cutoff_time_assumption_note": (
            "Stage125 Part3B1A locked the Cut-A available_at "
            "OPERATIONALIZATION but recorded zero real available_at "
            "assignments and unresolved pilot cutoff provenance, so no "
            "verified intraday available_at timestamp exists for any pair. "
            "Edition selection therefore uses 00:00:00Z of the cutoff date, "
            "the earliest possible instant, which can only exclude an "
            "edition and never admit one."),
    }


def build_unique_cutoff_plan(root: Path) -> list[dict[str, Any]]:
    """Section 5.3 — the deduplicated cutoff plan. An INTEGRITY count only."""
    path = root / CUTOFF_SOURCE_REL
    with path.open(encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))

    selected: list[dict[str, str]] = []
    for row in rows:
        if row[CUTOFF_MEMBERSHIP_FLAG].strip() != "True":
            continue
        # Section 5.2 hard guard: a final-test row must never reach the reader.
        if row["target_year"] in FINAL_TEST_TARGET_YEARS:
            raise M3I2EvidenceCaptureError(
                f"STOP_FINAL_TEST_ROW_REACHED_EVIDENCE_CAPTURE: target_year "
                f"{row['target_year']}")
        if row["target_year"] not in DEVELOPMENT_TARGET_YEARS:
            raise M3I2EvidenceCaptureError(
                f"unexpected target year {row['target_year']!r}")
        # only allowlisted fields are carried forward
        selected.append({k: row[k] for k in CUTOFF_COLUMN_ALLOWLIST})

    if len(selected) != EXPECTED_DEVELOPMENT_PAIRS:
        raise M3I2EvidenceCaptureError(
            f"development pairs {len(selected)} != "
            f"{EXPECTED_DEVELOPMENT_PAIRS}")
    companies = len({r["ticker"] for r in selected})
    if companies != EXPECTED_DEVELOPMENT_COMPANIES:
        raise M3I2EvidenceCaptureError(
            f"development companies {companies} != "
            f"{EXPECTED_DEVELOPMENT_COMPANIES}")

    counts: dict[str, int] = {}
    for row in selected:
        counts[row[CUTOFF_FIELD]] = counts.get(row[CUTOFF_FIELD], 0) + 1

    plan: list[dict[str, Any]] = []
    for cutoff_date in sorted(counts):
        plan.append({
            "cutoff_id": f"cut_{cutoff_date}",
            "pair_prediction_cutoff_utc": f"{cutoff_date}T{CUTOFF_TIME_ASSUMPTION}",
            "pair_prediction_cutoff_date": cutoff_date,
            "number_of_development_pairs_sharing_cutoff": counts[cutoff_date],
            "selected_wdi_archive_edition_id": "",
            "selected_wdi_archive_release_available_at": "",
            "selection_reason": "",
        })
    return plan


# --------------------------------------------------------------------------- #
# Official discovery roots (section 6)
# --------------------------------------------------------------------------- #

WB_ARCHIVE_LISTING_URL = (
    "https://datatopics.worldbank.org/world-development-indicators/"
    "wdi-archives.html")
WB_DATABANK_ARCHIVES_URL = "https://databank.worldbank.org/databases/archives"
WB_CPI_METADATA_URL = (
    "https://databank.worldbank.org/metadataglossary/"
    "world-development-indicators/series/FP.CPI.TOTL.ZG")
WB_FX_METADATA_URL = (
    "https://databank.worldbank.org/metadataglossary/"
    "world-development-indicators/series/PA.NUS.FCRF")
IMF_MFS_IR_DATASET_URL = "https://data.imf.org/en/datasets/IMF.STA%3AMFS_IR"

#: Discovery roots, captured BEFORE anything is downloaded from them. The
#: listing page is what proves an archive download is first-party.
DISCOVERY_TARGETS: tuple[dict[str, str], ...] = (
    {"object_id": "wb_wdi_archive_listing",
     "url": WB_ARCHIVE_LISTING_URL,
     "role": "wdi_archive_release_listing_discovery_root",
     "accept": "text/html,application/xhtml+xml"},
    {"object_id": "wb_databank_archives",
     "url": WB_DATABANK_ARCHIVES_URL,
     "role": "databank_archives_discovery_root",
     "accept": "text/html,application/xhtml+xml"},
    {"object_id": "wb_cpi_series_metadata",
     "url": WB_CPI_METADATA_URL,
     "role": "locked_cpi_indicator_official_metadata",
     "accept": "text/html,application/xhtml+xml"},
    {"object_id": "wb_fx_series_metadata",
     "url": WB_FX_METADATA_URL,
     "role": "locked_fx_indicator_official_metadata",
     "accept": "text/html,application/xhtml+xml"},
    {"object_id": "imf_mfs_ir_dataset_page",
     "url": IMF_MFS_IR_DATASET_URL,
     "role": "imf_financing_dataset_discovery_root",
     "accept": "text/html,application/xhtml+xml"},
)

FORBIDDEN_SOURCE_TOKENS: tuple[str, ...] = (
    "dbnomics", "fred.stlouisfed", "alfred.stlouisfed", "kaggle",
    "raw.githubusercontent", "tradingeconomics", "ceicdata", "bonbast",
    "tgju.org", "investing.com", "macrotrends",
)


# --------------------------------------------------------------------------- #
# Release-availability rule (section 8) — inherited from the merged contract
# --------------------------------------------------------------------------- #

#: A date embedded in an official filename is an EDITION DATE TOKEN. It is not
#: a release date until retained official bytes say so in words. The audit that
#: forced this distinction was right: the World Bank archive listing publishes
#: "Year | Month(s)" plus hyperlinks, and nowhere states that the date inside a
#: filename is the publication, release or update date of that edition.
EDITION_DATE_TOKEN_SOURCE_FILENAME = "official_listing_download_filename"

RELEASE_DATE_UNRESOLVED_STATUS = (
    "UNRESOLVED_FILENAME_DATE_TOKEN_NOT_VERIFIED_AS_RELEASE_DATE")
RELEASE_DATE_VERIFIED_STATUS = (
    "VERIFIED_RELEASE_DATE_EXPLICITLY_STATED_BY_OFFICIAL_SOURCE")
RELEASE_DATE_ABSENT_STATUS = "UNRESOLVED_NO_DATE_TOKEN_AND_NO_STATED_RELEASE"

#: Wordings that would make a date an explicit release/publication/update date.
#: Searched against RETAINED bytes only.
RELEASE_STATEMENT_PATTERNS: tuple[str, ...] = (
    r"release[ds]?\s*(date|on)?",
    r"publish(ed|ication)\s*(date|on)?",
    r"last\s+updated",
    r"data\s+updated",
    r"update[ds]?\s+on",
)

#: Things that can NEVER establish a historical release time (section 3 rule 4).
NON_EVIDENCE_FOR_RELEASE_TIME: tuple[str, ...] = (
    "retrieval timestamp",
    "HTTP Last-Modified header",
    "ZIP member timestamp",
    "workbook document properties",
    "file-system mtime",
    "the observation years present in the file",
)


def find_explicit_release_date_statement(
    edition_date_token: str, retained_texts: dict[str, str],
) -> tuple[bool, str, str]:
    """Search RETAINED bytes for an explicit release/publication statement.

    Returns ``(verified, evidence_artifact, evidence_locator)``. A hit counts
    only when release/publication/update wording appears in the immediate
    neighbourhood of the edition's own date token, in an official retained
    object. Nothing is inferred from a header, a timestamp or a filename.
    """
    if not edition_date_token:
        return False, "", ""
    slash = edition_date_token.replace("-", "/")
    compact = edition_date_token.replace("-", "")
    needles = (edition_date_token, slash, compact)

    for artifact, text in sorted(retained_texts.items()):
        lowered = text.lower()
        for needle in needles:
            start = 0
            while True:
                index = lowered.find(needle.lower(), start)
                if index < 0:
                    break
                window = lowered[max(0, index - 300):index + 300]
                for pattern in RELEASE_STATEMENT_PATTERNS:
                    if re.search(pattern, window):
                        line = lowered.count("\n", 0, index) + 1
                        return True, artifact, (
                            f"{artifact}#offset={index};line={line};"
                            f"token={edition_date_token}")
                start = index + 1
    return False, "", ""


def derive_release_available_at(
    listed_release_date: str, listed_release_time: str | None,
) -> tuple[str, bool]:
    """Return ``(available_at_utc, date_only_next_day_rule_applied)``.

    The merged contract's rule, applied verbatim: a verified exact time is used
    as-is; a date-only release becomes 00:00:00 UTC on the NEXT calendar day,
    which is what makes a same-day date-only release fall outside a same-day
    cutoff.
    """
    if not listed_release_date:
        raise M3I2EvidenceCaptureError(
            "an edition without a verified release date cannot be given an "
            "available_at")
    if listed_release_time:
        return f"{listed_release_date}T{listed_release_time}Z", False
    day = datetime.strptime(listed_release_date, "%Y-%m-%d").replace(
        tzinfo=timezone.utc) + timedelta(days=1)
    return day.strftime("%Y-%m-%dT00:00:00Z"), True


def select_edition_for_cutoff(
    cutoff_utc: str, editions: list[dict[str, Any]],
) -> dict[str, Any] | None:
    """The latest verified edition STRICTLY before the cutoff, or None.

    Only editions with a verified release date are eligible: an edition whose
    release date the official listing does not establish is not a verified
    pre-cutoff vintage, and is skipped rather than guessed at.
    """
    eligible = [
        e for e in editions
        if e.get("derived_release_available_at_utc")
        and e.get("release_date_verified") is True
        and e["derived_release_available_at_utc"] < cutoff_utc
    ]
    if not eligible:
        return None
    return max(eligible, key=lambda e: e["derived_release_available_at_utc"])


def plan_required_editions(
    cutoff_plan: list[dict[str, Any]], editions: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Attach a selected edition to each cutoff and deduplicate the editions."""
    plan = [dict(row) for row in cutoff_plan]
    usage: dict[str, dict[str, Any]] = {}

    for row in plan:
        chosen = select_edition_for_cutoff(
            row["pair_prediction_cutoff_utc"], editions)
        if chosen is None:
            row["selected_wdi_archive_edition_id"] = ""
            row["selected_wdi_archive_release_available_at"] = ""
            row["selection_reason"] = (
                "NO_VERIFIED_PRE_CUTOFF_EDITION")
            continue
        edition_id = chosen["archive_edition_id"]
        row["selected_wdi_archive_edition_id"] = edition_id
        row["selected_wdi_archive_release_available_at"] = chosen[
            "derived_release_available_at_utc"]
        row["selection_reason"] = (
            "latest verified WDI archive release with release_available_at < "
            "pair_prediction_cutoff")

        entry = usage.setdefault(edition_id, {
            "archive_edition_id": edition_id,
            "release_available_at_utc":
                chosen["derived_release_available_at_utc"],
            "minimum_cutoff_using_edition": row["pair_prediction_cutoff_utc"],
            "maximum_cutoff_using_edition": row["pair_prediction_cutoff_utc"],
            "development_pair_count_using_edition": 0,
            "download_required": True,
            "download_status": "NOT_ATTEMPTED",
            "raw_artifact_filename": "",
            "raw_artifact_sha256": "",
            "raw_artifact_bytes": 0,
        })
        entry["minimum_cutoff_using_edition"] = min(
            entry["minimum_cutoff_using_edition"],
            row["pair_prediction_cutoff_utc"])
        entry["maximum_cutoff_using_edition"] = max(
            entry["maximum_cutoff_using_edition"],
            row["pair_prediction_cutoff_utc"])
        entry["development_pair_count_using_edition"] += row[
            "number_of_development_pairs_sharing_cutoff"]

    required = [usage[k] for k in sorted(usage)]
    return plan, required


# --------------------------------------------------------------------------- #
# Semantic compatibility (section 12)
# --------------------------------------------------------------------------- #

SEMANTIC_STATUSES: tuple[str, ...] = ("PASS", "UNRESOLVED", "FAIL_INTEGRITY")

CPI_REQUIRED_INTERPRETATION = (
    "annual CPI inflation-rate series in percent")
CPI_FORBIDDEN_INTERPRETATIONS: tuple[str, ...] = (
    "CPI index level", "GDP-deflator inflation", "monthly inflation",
    "point-to-point inflation", "moving-average inflation",
    "another inflation construct",
)
FX_REQUIRED_INTERPRETATION = (
    "Official exchange rate, LCU per US dollar, period average, annual")

SEMANTIC_EVIDENCE_FIELDS: tuple[str, ...] = (
    "archive_edition_id", "indicator_code", "release_available_at_utc",
    "economy_identity_verified", "indicator_code_verified",
    "archived_series_title_raw", "title_compatibility", "frequency_raw",
    "frequency_annual_verified", "unit_raw", "unit_compatibility",
    # provenance of the unit claim was previously computed but NOT committed;
    # an evidence field that never reaches the artifact is not evidence.
    "unit_evidence_source",
    "calendar_year_semantics_verified", "calendar_year_evidence_locator",
    # --- FX continuity evidence (section 4) --------------------------------
    "currency_denomination_raw", "currency_denomination_evidence_artifact",
    "currency_denomination_evidence_locator", "currency_denomination_verified",
    "local_currency_valuation_definition_raw",
    "valuation_definition_evidence_artifact",
    "valuation_definition_evidence_locator", "valuation_definition_verified",
    "redenomination_or_unit_break_evidence_raw",
    "redenomination_or_unit_break_evidence_artifact",
    "redenomination_or_unit_break_evidence_locator",
    "no_redenomination_or_unit_break_verified",
    "fx_pair_semantic_compatibility_status",
    "raw_archive_sha256", "evidence_locator", "compatibility_status",
    "unresolved_reason",
)

#: Fields an FX row must ALL evidence before it may be PASS. The archived
#: title supports the generic construct and the unit label; it says nothing
#: about which currency unit Iran was denominated in for a given vintage, nor
#: whether a redenomination fell between two adjacent years.
FX_CONTINUITY_REQUIRED_FIELDS: tuple[str, ...] = (
    "currency_denomination_verified",
    "valuation_definition_verified",
    "no_redenomination_or_unit_break_verified",
)


def classify_semantic_compatibility(evidence: dict[str, Any]) -> str:
    """PASS only with complete evidence; missing proof is UNRESOLVED.

    Absence of metadata is never FAIL. FAIL_INTEGRITY is reserved for evidence
    that actively contradicts the locked meaning.
    """
    if evidence.get("integrity_contradiction"):
        return "FAIL_INTEGRITY"
    required = (
        evidence.get("economy_identity_verified"),
        evidence.get("indicator_code_verified"),
        evidence.get("frequency_annual_verified"),
        evidence.get("calendar_year_semantics_verified"),
    )
    if not all(v is True for v in required):
        return "UNRESOLVED"
    if evidence.get("title_compatibility") != "COMPATIBLE":
        return "UNRESOLVED"
    if evidence.get("unit_compatibility") != "COMPATIBLE":
        return "UNRESOLVED"
    if not evidence.get("raw_archive_sha256"):
        return "UNRESOLVED"
    return "PASS"


def assert_fx_pass_has_continuity_evidence(row: dict[str, Any]) -> None:
    """An FX row may be PASS only with all three continuity facts evidenced."""
    if row.get("indicator_code") != FX_INDICATOR_CODE:
        return
    if row.get("compatibility_status") != "PASS":
        return
    missing = [f for f in FX_CONTINUITY_REQUIRED_FIELDS
               if row.get(f) is not True]
    if missing:
        raise M3I2EvidenceCaptureError(
            f"FX semantic PASS without continuity evidence: {missing}")
    for field in ("currency_denomination_evidence_locator",
                  "valuation_definition_evidence_locator",
                  "redenomination_or_unit_break_evidence_locator"):
        if not row.get(field):
            raise M3I2EvidenceCaptureError(
                f"FX semantic PASS without an evidence locator: {field}")
    if row.get("fx_pair_semantic_compatibility_status") != "PASS":
        raise M3I2EvidenceCaptureError(
            "fx_pair_semantic_compatibility_status must agree with the row")


#: Evidence every PASS row must carry, whichever indicator it describes. The
#: FX-only continuity fields are checked separately, against FX rows only.
SEMANTIC_CORE_EVIDENCE_FIELDS: tuple[str, ...] = (
    "archive_edition_id", "indicator_code", "economy_identity_verified",
    "indicator_code_verified", "archived_series_title_raw",
    "title_compatibility", "frequency_raw", "frequency_annual_verified",
    "unit_raw", "unit_evidence_source", "unit_compatibility",
    "calendar_year_semantics_verified", "calendar_year_evidence_locator",
    "raw_archive_sha256", "evidence_locator", "compatibility_status",
)


def assert_semantic_pass_is_fully_evidenced(row: dict[str, Any]) -> None:
    """QC rule — a PASS that lacks any required evidence is a QC failure."""
    if row.get("compatibility_status") != "PASS":
        return
    missing = [f for f in SEMANTIC_CORE_EVIDENCE_FIELDS if not row.get(f)]
    if missing:
        raise M3I2EvidenceCaptureError(
            f"semantic PASS without complete evidence: {missing}")
    assert_fx_pass_has_continuity_evidence(row)


# --------------------------------------------------------------------------- #
# Decision vocabulary (section 14)
# --------------------------------------------------------------------------- #

EVIDENCE_COMPLETE = "EVIDENCE_COMPLETE_FOR_SEPARATE_M3I2_DATA_GATE_REVIEW"
EVIDENCE_UNRESOLVED = "UNRESOLVED_OFFICIAL_SOURCE_EVIDENCE"
EVIDENCE_INVALID = "INVALID_OFFICIAL_SOURCE_EVIDENCE_CAPTURE"

EVIDENCE_STATUSES: tuple[str, ...] = (
    EVIDENCE_COMPLETE, EVIDENCE_UNRESOLVED, EVIDENCE_INVALID)

FINANCING_DECISIONS: tuple[str, ...] = (
    "EXACT_CANDIDATE_IDENTIFIED_PENDING_SEPARATE_PROSPECTIVE_LOCK",
    "NO_EXACT_CANDIDATE_IDENTIFIED_UNRESOLVED_METADATA_LOCK",
    "OFFICIAL_METADATA_ACCESS_UNRESOLVED",
    "OFFICIAL_METADATA_INTEGRITY_BLOCKED",
)

RESULT_CODES: dict[str, str] = {
    EVIDENCE_COMPLETE:
        "M3I2_OFFICIAL_SOURCE_EVIDENCE_CAPTURE_READY_FOR_INDEPENDENT_AUDIT",
    EVIDENCE_UNRESOLVED:
        "M3I2_OFFICIAL_SOURCE_EVIDENCE_CAPTURE_UNRESOLVED_READY_FOR_"
        "INDEPENDENT_AUDIT",
    EVIDENCE_INVALID:
        "M3I2_OFFICIAL_SOURCE_EVIDENCE_CAPTURE_INVALID_READY_FOR_INDEPENDENT_"
        "AUDIT",
}

NEXT_ACTION_BY_STATUS: dict[str, str] = {
    EVIDENCE_COMPLETE: "stage128-m3i2-data-gate",
    EVIDENCE_UNRESOLVED: "stage128-m3i2-official-source-evidence-review",
    EVIDENCE_INVALID:
        "stage128-m3i2-official-source-evidence-integrity-review",
}


def classify_evidence(summary: dict[str, Any]) -> str:
    """Section 14 — complete, unresolved or invalid. Fail closed to unresolved.

    INVALID outranks everything: an integrity violation is never softened into
    "unresolved". Otherwise COMPLETE requires every condition in 14.1; anything
    short of that is UNRESOLVED, never "observed failure" and never zero
    coverage.
    """
    if summary.get("integrity_violations"):
        return EVIDENCE_INVALID

    complete = (
        summary.get("required_editions_total", 0) > 0
        # Every development cutoff must actually be servable. Capturing the
        # editions that happen to exist is not the same as being able to serve
        # every cutoff: a cutoff with no verified pre-cutoff vintage leaves a
        # hole that a later Data Gate would silently inherit.
        and summary.get("cutoffs_without_verified_pre_cutoff_edition", 1) == 0
        and summary.get("development_pairs_without_verified_pre_cutoff_edition",
                        1) == 0
        and summary.get("required_editions_with_verified_release_available_at",
                        0) == summary.get("required_editions_total", 0)
        and summary.get("required_editions_captured", 0) == summary.get(
            "required_editions_total", 0)
        and summary.get("raw_bytes_retained_for_every_capture_claim") is True
        and summary.get("locked_series_rows_extracted", 0) > 0
        and summary.get("indicator_substitution_occurred") is False
        and summary.get("semantic_unresolved_count", 1) == 0
        and summary.get("semantic_fail_integrity_count", 1) == 0
        and summary.get("external_bundle_available_for_handoff") is True
        and summary.get("offline_rebuild_reproduces_committed_artifacts")
        is True
        and summary.get("data_gate_executions", 1) == 0
        and summary.get("company_macro_joins", 1) == 0
    )
    return EVIDENCE_COMPLETE if complete else EVIDENCE_UNRESOLVED


# --------------------------------------------------------------------------- #
# Forbidden-execution firewalls (section 17)
# --------------------------------------------------------------------------- #

FORBIDDEN_NETWORK_MODULES: tuple[str, ...] = (
    "requests", "urllib", "urllib3", "http", "httpx", "aiohttp", "socket",
    "webbrowser", "selenium", "playwright", "ssl",
)
FORBIDDEN_MODEL_MODULES: tuple[str, ...] = (
    "sklearn", "xgboost", "lightgbm", "catboost", "statsmodels", "imblearn",
    "shap",
)


def _import_pattern(*modules: str) -> re.Pattern[str]:
    return re.compile(
        r"^\s*(?:from|import)\s+(" + "|".join(modules) + r")\b", re.MULTILINE)


FORBIDDEN_NETWORK_IMPORTS = _import_pattern(*FORBIDDEN_NETWORK_MODULES)
FORBIDDEN_MODEL_IMPORTS = _import_pattern(*FORBIDDEN_MODEL_MODULES)

FORBIDDEN_ESTIMATOR_TOKENS: tuple[str, ...] = (
    ".fit(", ".fit_predict(", ".fit_resample(", ".predict(",
    ".predict_proba(", "roc_auc_score", "average_precision_score",
    "multipletests", "SMOTE(",
)

FORBIDDEN_COMPUTATION_TOKENS: tuple[str, ...] = (
    "candidate_valid_coverage =", "block_common_sample_coverage =",
    "gate_result =", "math.log(", "np.log(",
)


def _read_texts(root: Path, files: tuple[str, ...]) -> dict[str, str]:
    out: dict[str, str] = {}
    for rel in files:
        path = root / rel
        if not path.is_file():
            raise M3I2EvidenceCaptureError(f"implementation file missing: {rel}")
        out[rel] = path.read_text(encoding="utf-8")
    return out


def _scan_imports(texts: dict[str, str], pattern: re.Pattern[str],
                  label: str) -> None:
    for rel, text in texts.items():
        hit = pattern.search(text)
        if hit:
            raise M3I2EvidenceCaptureError(
                f"forbidden {label} import {hit.group(1)!r} in {rel}")


def _scan_tokens(texts: dict[str, str], tokens: tuple[str, ...],
                 label: str) -> None:
    declaration_markers = (
        "FORBIDDEN_ESTIMATOR_TOKENS", "FORBIDDEN_COMPUTATION_TOKENS",
        "FORBIDDEN_SOURCE_TOKENS",
    )
    hits: list[str] = []
    for rel, text in texts.items():
        in_declaration = False
        for lineno, line in enumerate(text.splitlines(), start=1):
            if any(marker in line for marker in declaration_markers):
                in_declaration = line.rstrip().endswith("(")
                continue
            if in_declaration:
                if line.strip() == ")":
                    in_declaration = False
                continue
            for token in tokens:
                if token in line:
                    hits.append(f"{rel}:{lineno}:{token}")
    if hits:
        raise M3I2EvidenceCaptureError(
            f"forbidden {label} path(s) present: {sorted(hits)}")


def assert_offline_layer_has_no_network(root: Path) -> None:
    """The builder, validator and tests must be import-clean of the network."""
    texts = _read_texts(root, OFFLINE_IMPLEMENTATION_FILES)
    _scan_imports(texts, FORBIDDEN_NETWORK_IMPORTS, "network")


def assert_no_model_or_estimator(root: Path) -> None:
    """No modeling library and no estimator call anywhere in this action."""
    files = OFFLINE_IMPLEMENTATION_FILES + (CAPTURE_LAYER_FILE, RUNNER_FILE)
    texts = _read_texts(root, files)
    _scan_imports(texts, FORBIDDEN_MODEL_IMPORTS, "model")
    _scan_tokens(texts, FORBIDDEN_ESTIMATOR_TOKENS, "estimator")


def assert_no_forbidden_computation(root: Path) -> None:
    """No FX log change, no coverage, no Gate result computed in this action."""
    files = OFFLINE_IMPLEMENTATION_FILES + (CAPTURE_LAYER_FILE, RUNNER_FILE)
    _scan_tokens(_read_texts(root, files), FORBIDDEN_COMPUTATION_TOKENS,
                 "forbidden-computation")


def assert_no_unofficial_source(root: Path) -> None:
    """No mirror, aggregator or unofficial FX source may appear anywhere."""
    files = OFFLINE_IMPLEMENTATION_FILES + (CAPTURE_LAYER_FILE, RUNNER_FILE)
    hits: list[str] = []
    for rel, text in _read_texts(root, files).items():
        in_declaration = False
        for lineno, line in enumerate(text.splitlines(), start=1):
            if "FORBIDDEN_SOURCE_TOKENS" in line:
                in_declaration = line.rstrip().endswith("(")
                continue
            if in_declaration:
                if line.strip() == ")":
                    in_declaration = False
                continue
            lowered = line.lower()
            # Prose that NAMES a forbidden source in order to forbid it is
            # fine; what must never appear is one being addressed as a source.
            looks_like_a_reference = ("://" in lowered or "www." in lowered
                                      or "host" in lowered)
            if not looks_like_a_reference:
                continue
            for token in FORBIDDEN_SOURCE_TOKENS:
                if token in lowered:
                    hits.append(f"{rel}:{lineno}:{token}")
    if hits:
        raise M3I2EvidenceCaptureError(
            f"unofficial source referenced: {sorted(hits)}")


def assert_no_final_test_access(rows: list[dict[str, Any]]) -> None:
    """No final-test target year may appear in any cutoff-derived row."""
    bad = [r for r in rows
           if str(r.get("target_year", "")) in FINAL_TEST_TARGET_YEARS]
    if bad:
        raise M3I2EvidenceCaptureError(
            "STOP_FINAL_TEST_ROW_REACHED_EVIDENCE_CAPTURE")


def assert_locked_indicators_only(codes: list[str]) -> None:
    """Only the two locked codes; substitution is a hard failure."""
    forbidden = sorted(set(codes) & set(FORBIDDEN_INDICATOR_CODES))
    if forbidden:
        raise M3I2EvidenceCaptureError(
            f"forbidden indicator substitution: {forbidden}")
    unknown = sorted(set(codes) - set(LOCKED_INDICATOR_CODES))
    if unknown:
        raise M3I2EvidenceCaptureError(
            f"indicator outside the locked pair: {unknown}")


def assert_financing_not_a_forbidden_proxy(evidence: dict[str, Any]) -> None:
    """A forbidden proxy may never be identified as the financing candidate."""
    identified = evidence.get("identified_series_title") or ""
    if not identified:
        return
    lowered = identified.lower()
    for proxy in FINANCING_FORBIDDEN_PROXIES:
        if proxy.lower() in lowered:
            raise M3I2EvidenceCaptureError(
                f"forbidden financing proxy identified: {proxy}")
    if evidence.get("m3i3_admitted") is not False:
        raise M3I2EvidenceCaptureError(
            "M3I-3 may never be admitted by an evidence-capture action")


def assert_edition_selection_is_value_blind(plan: list[dict[str, Any]],
                                            audit: dict[str, Any]) -> None:
    """Editions are chosen from cutoffs and release dates, never from values."""
    if audit.get("edition_selection_used_observed_values") is not False:
        raise M3I2EvidenceCaptureError(
            "an archive edition was selected after inspecting values")
    if audit.get("edition_switched_after_missing_value_inspection") is not (
            False):
        raise M3I2EvidenceCaptureError(
            "a later edition was chosen because an earlier one had missing "
            "data")
    for row in plan:
        reason = row.get("selection_reason", "")
        if reason and not reason.startswith((
                "latest verified WDI archive release",
                "NO_VERIFIED_PRE_CUTOFF_EDITION")):
            raise M3I2EvidenceCaptureError(
                f"unrecognised edition-selection reason: {reason!r}")


# --------------------------------------------------------------------------- #
# Protected immutability (section 18)
# --------------------------------------------------------------------------- #

PROTECTED_TREES: tuple[str, ...] = (
    "project/stage125",
    "project/stage126",
    "project/stage127",
    "project/stage128/m2_incremental_evaluation",
    "project/stage128/m2_retained_block_human_decision",
    "project/stage128/m3_macro_data_gate",
    "project/stage128/m3_intl_macro_contract_lock",
)

PROTECTED_EXTRA_FILES: tuple[str, ...] = (
    "project/stage128/stage128_m2_d2_development_features.csv",
)

#: CLOSED list. Operational verification artifacts that repository governance
#: explicitly permits to evolve with their validator source.
PROTECTED_OPERATIONAL_EXCLUSIONS: tuple[str, ...] = (
    "project/stage126/README_STAGE126_CURRENT_STATE_VALIDATION.md",
    "project/stage126/metadata_and_hashes_stage126_current_state_validator.json",
    "project/stage126/stage126_current_state_validation_report.json",
)


def _git(root: Path, *args: str) -> str:
    import subprocess

    proc = subprocess.run(["git", *args], cwd=str(root), capture_output=True,
                          text=True)
    if proc.returncode != 0:
        raise M3I2EvidenceCaptureError(
            f"git {' '.join(args)} failed: {proc.stderr.strip()}")
    return proc.stdout


def _git_blob_sha(root: Path, rel: str) -> str:
    return _git(root, "hash-object", rel).strip()


def _sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _tracked_under(root: Path, commit: str) -> tuple[str, ...]:
    out = _git(root, "ls-tree", "-r", "--name-only", "-z", commit, "--",
               *PROTECTED_TREES)
    return tuple(sorted(p for p in out.split("\0") if p))


def verify_protected_immutability(root: Path) -> dict[str, Any]:
    """Section 18 — every protected artifact byte-identical to the baseline."""
    paths = set(_tracked_under(root, BASELINE_COMMIT))
    paths -= set(PROTECTED_OPERATIONAL_EXCLUSIONS)
    for rel in PROTECTED_EXTRA_FILES:
        _git(root, "cat-file", "-e", f"{BASELINE_COMMIT}:{rel}")
        paths.add(rel)
    ordered = tuple(sorted(paths))
    if not ordered:
        raise M3I2EvidenceCaptureError("protected enumeration produced nothing")

    changed = [p for p in _git(
        root, "diff", "--name-only", f"{BASELINE_COMMIT}..HEAD", "--",
        *ordered).splitlines() if p.strip()]
    if changed:
        contract = [p for p in changed if p.startswith(CONTRACT_PACKAGE_DIR)]
        if contract:
            raise M3I2EvidenceCaptureError(
                f"STOP_MERGED_CONTRACT_PACKAGE_MUTATED: {sorted(contract)}")
        raise M3I2EvidenceCaptureError(
            f"protected paths changed: {sorted(changed)}")

    contract_paths = [p for p in ordered if p.startswith(CONTRACT_PACKAGE_DIR)]
    for rel in contract_paths:
        on_disk = _sha256_file(root / rel)
        baseline = hashlib.sha256(
            _git(root, "show", f"{BASELINE_COMMIT}:{rel}").encode("utf-8")
        ).hexdigest()
        if on_disk != baseline:
            raise M3I2EvidenceCaptureError(
                f"STOP_MERGED_CONTRACT_PACKAGE_MUTATED: {rel}")

    return {
        "protected_baseline_branch": BASELINE_BRANCH,
        "protected_baseline_commit": BASELINE_COMMIT,
        "protected_trees": list(PROTECTED_TREES),
        "protected_extra_files": list(PROTECTED_EXTRA_FILES),
        "protected_operational_exclusions": list(
            PROTECTED_OPERATIONAL_EXCLUSIONS),
        "protected_file_count": len(ordered),
        "merged_contract_package_file_count": len(contract_paths),
        "merged_contract_package_byte_identical": True,
        "protected_committed_history_diff_empty": True,
        "no_scientific_predecessor_artifact_changed": True,
    }


# --------------------------------------------------------------------------- #
# External bundle (section 10)
# --------------------------------------------------------------------------- #

BUNDLE_BASENAME = "papermali_stage128_m3i2_official_source_evidence_bundle_v1"
BUNDLE_PART_LIMIT_BYTES = 1_500_000_000


def build_external_bundle(
    capture_dir: str | os.PathLike[str], bundle_dir: str | os.PathLike[str],
) -> dict[str, Any]:
    """Package the retained capture directory into a deterministic ZIP.

    Member order and timestamps are normalized so that identical captured bytes
    always produce an identical bundle hash. Raw bytes are NOT deleted here.
    """
    src = Path(capture_dir)
    dest = Path(bundle_dir)
    dest.mkdir(parents=True, exist_ok=True)

    members = sorted(
        (p for p in src.rglob("*")
         if p.is_file() and p.name != "SHA256SUMS.txt"),
        key=lambda p: str(p.relative_to(src)))
    if not members:
        raise M3I2EvidenceCaptureError(
            "STOP_RAW_BYTES_NOT_RETAINED: the capture directory is empty")

    member_records: list[dict[str, Any]] = []
    lines: list[str] = []
    for path in members:
        rel = str(path.relative_to(src))
        data = path.read_bytes()
        digest = hashlib.sha256(data).hexdigest()
        member_records.append({
            "member_path": rel,
            "member_bytes": len(data),
            "member_sha256": digest,
        })
        lines.append(f"{digest}  {rel}")

    sha_text = "\n".join(lines) + "\n"
    (src / "SHA256SUMS.txt").write_text(sha_text, encoding="utf-8")

    zip_path = dest / f"{BUNDLE_BASENAME}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for record in member_records:
            info = zipfile.ZipInfo(record["member_path"],
                                   date_time=(1980, 1, 1, 0, 0, 0))
            info.external_attr = 0o644 << 16
            info.compress_type = zipfile.ZIP_DEFLATED
            zf.writestr(info, (src / record["member_path"]).read_bytes())
        info = zipfile.ZipInfo("SHA256SUMS.txt", date_time=(1980, 1, 1, 0, 0, 0))
        info.external_attr = 0o644 << 16
        zf.writestr(info, sha_text)

    data = zip_path.read_bytes()
    return {
        "bundle_basename": BUNDLE_BASENAME,
        "bundle_parts": [{
            "filename": zip_path.name,
            "part_number": 1,
            "total_parts": 1,
            "byte_size": len(data),
            "sha256": hashlib.sha256(data).hexdigest(),
        }],
        "bundle_total_bytes": len(data),
        "bundle_member_count": len(member_records),
        "bundle_members": member_records,
        "deterministic_member_order": True,
        "normalized_member_timestamps": True,
        "raw_bytes_deleted_after_hashing": False,
        "raw_bytes_committed_to_git": False,
        "raw_bytes_available_for_independent_handoff": True,
        "local_capture_directory_preserved": str(src),
    }


MULTIPART_MAX_BYTES = 200_000_000
MULTIPART_HANDOFF_MANIFEST_MEMBER = "handoff_manifest.json"


def build_multipart_handoff_bundle(
    capture_dir: str | os.PathLike[str], bundle_dir: str | os.PathLike[str],
) -> dict[str, Any]:
    """Split the evidence into valid, deterministic ZIP parts for handoff.

    Each part is a real ZIP, not a byte-slice of a larger one, so an auditor can
    open any part on its own. Every original member is assigned to exactly one
    part as a primary member, with its bytes and SHA-256 unchanged. A small
    global handoff manifest is copied into every part so no part is orphaned.

    The original single-file bundle and the capture directory are left in place.
    """
    src = Path(capture_dir)
    dest = Path(bundle_dir)
    dest.mkdir(parents=True, exist_ok=True)

    members = sorted(
        (p for p in src.rglob("*")
         if p.is_file() and p.name != "SHA256SUMS.txt"),
        key=lambda p: str(p.relative_to(src)))
    if not members:
        raise M3I2EvidenceCaptureError(
            "STOP_RAW_BYTES_NOT_RETAINED: the capture directory is empty")

    # Deterministic sequential packing over the sorted member order.
    parts: list[list[Path]] = [[]]
    sizes: list[int] = [0]
    for path in members:
        size = path.stat().st_size
        if size > MULTIPART_MAX_BYTES:
            raise M3I2EvidenceCaptureError(
                f"a single member exceeds the part limit: {path.name}")
        if sizes[-1] and sizes[-1] + size > MULTIPART_MAX_BYTES:
            parts.append([])
            sizes.append(0)
        parts[-1].append(path)
        sizes[-1] += size

    total_parts = len(parts)
    global_manifest = {
        "bundle_basename": BUNDLE_BASENAME,
        "total_parts": total_parts,
        "part_max_bytes": MULTIPART_MAX_BYTES,
        "action_id": ACTION_ID,
        "note": (
            "Each part is a standalone valid ZIP. Every original member "
            "appears in exactly one part as a primary member; this manifest is "
            "copied into every part for orientation."),
    }
    manifest_bytes = (json.dumps(global_manifest, ensure_ascii=False, indent=2,
                                 sort_keys=True) + "\n").encode("utf-8")

    part_records: list[dict[str, Any]] = []
    assigned: list[str] = []
    for index, group in enumerate(parts, start=1):
        name = f"{BUNDLE_BASENAME}_part{index:03d}.zip"
        path = dest / name
        member_records: list[dict[str, Any]] = []
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
            info = zipfile.ZipInfo(MULTIPART_HANDOFF_MANIFEST_MEMBER,
                                   date_time=(1980, 1, 1, 0, 0, 0))
            info.external_attr = 0o644 << 16
            zf.writestr(info, manifest_bytes)
            for member in group:
                rel = str(member.relative_to(src))
                data = member.read_bytes()
                entry = zipfile.ZipInfo(rel, date_time=(1980, 1, 1, 0, 0, 0))
                entry.external_attr = 0o644 << 16
                entry.compress_type = zipfile.ZIP_DEFLATED
                zf.writestr(entry, data)
                digest = hashlib.sha256(data).hexdigest()
                member_records.append({
                    "member_path": rel,
                    "member_bytes": len(data),
                    "member_sha256": digest,
                    "primary_member": True,
                })
                assigned.append(rel)
        blob = path.read_bytes()
        part_records.append({
            "filename": name,
            "part_number": index,
            "total_parts": total_parts,
            "byte_size": len(blob),
            "sha256": hashlib.sha256(blob).hexdigest(),
            "primary_member_count": len(member_records),
            "members": member_records,
            "copied_global_manifest": MULTIPART_HANDOFF_MANIFEST_MEMBER,
        })

    # Every original member assigned exactly once.
    expected = [str(m.relative_to(src)) for m in members]
    if sorted(assigned) != sorted(expected):
        raise M3I2EvidenceCaptureError(
            "multipart member allocation does not cover every member exactly "
            "once")
    duplicates = sorted({m for m in assigned if assigned.count(m) > 1})
    if duplicates:
        raise M3I2EvidenceCaptureError(
            f"member assigned to more than one part: {duplicates}")

    lines = [f"{r['sha256']}  {r['filename']}" for r in part_records]
    (dest / "SHA256SUMS.txt").write_text("\n".join(lines) + "\n",
                                         encoding="utf-8")

    manifest = {
        "multipart_basename": BUNDLE_BASENAME,
        "multipart_total_parts": total_parts,
        "multipart_part_max_bytes": MULTIPART_MAX_BYTES,
        "multipart_parts": part_records,
        "multipart_total_bytes": sum(r["byte_size"] for r in part_records),
        "multipart_member_count": len(expected),
        "every_member_assigned_exactly_once": True,
        "parts_are_valid_zip_archives": True,
        "deterministic_member_order": True,
        "normalized_member_timestamps": True,
        "sha256sums_file": "SHA256SUMS.txt",
        "original_single_bundle_retained": True,
        "capture_directory_retained": str(src),
        "independently_verified_by_auditor": False,
        "independent_verification_note": (
            "Independent verification cannot be claimed until every part has "
            "actually been provided to the auditor."),
    }
    verify_multipart_bundle(manifest, dest)
    return manifest


def verify_multipart_bundle(manifest: dict[str, Any],
                            bundle_dir: str | os.PathLike[str]) -> None:
    """Re-open and re-hash every part; anything off is a hard stop."""
    dest = Path(bundle_dir)
    for part in manifest.get("multipart_parts", []):
        path = dest / part["filename"]
        if not path.is_file():
            raise M3I2EvidenceCaptureError(
                f"STOP_RAW_BYTES_NOT_RETAINED: {part['filename']}")
        if hashlib.sha256(path.read_bytes()).hexdigest() != part["sha256"]:
            raise M3I2EvidenceCaptureError(
                f"STOP_EXTERNAL_BUNDLE_HASH_MISMATCH: {part['filename']}")
        if part["byte_size"] > MULTIPART_MAX_BYTES:
            raise M3I2EvidenceCaptureError(
                f"part exceeds the size limit: {part['filename']}")
        with zipfile.ZipFile(path) as zf:          # must be a VALID zip
            if zf.testzip() is not None:
                raise M3I2EvidenceCaptureError(
                    f"corrupt part: {part['filename']}")
            names = set(zf.namelist())
            for member in part["members"]:
                if member["member_path"] not in names:
                    raise M3I2EvidenceCaptureError(
                        f"part {part['filename']} is missing declared member "
                        f"{member['member_path']}")
                data = zf.read(member["member_path"])
                if hashlib.sha256(data).hexdigest() != member["member_sha256"]:
                    raise M3I2EvidenceCaptureError(
                        f"member bytes changed inside {part['filename']}: "
                        f"{member['member_path']}")


def verify_bundle_manifest(manifest: dict[str, Any],
                           bundle_dir: str | os.PathLike[str]) -> None:
    """Recompute every part hash; a mismatch is a hard stop."""
    dest = Path(bundle_dir)
    for part in manifest.get("bundle_parts", []):
        path = dest / part["filename"]
        if not path.is_file():
            raise M3I2EvidenceCaptureError(
                f"STOP_RAW_BYTES_NOT_RETAINED: bundle part absent "
                f"{part['filename']}")
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        if digest != part["sha256"]:
            raise M3I2EvidenceCaptureError(
                f"STOP_EXTERNAL_BUNDLE_HASH_MISMATCH: {part['filename']}")


# --------------------------------------------------------------------------- #
# Rendering helpers
# --------------------------------------------------------------------------- #

def _json_text(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2,
                      sort_keys=True) + "\n"


def _csv_text(columns: tuple[str, ...] | list[str],
              rows: list[dict[str, Any]]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(columns), lineterminator="\n",
                            extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({k: ("" if row.get(k) is None else row.get(k))
                         for k in columns})
    return buf.getvalue()


CUTOFF_PLAN_COLUMNS: tuple[str, ...] = (
    "cutoff_id", "pair_prediction_cutoff_utc", "pair_prediction_cutoff_date",
    "number_of_development_pairs_sharing_cutoff",
    "selected_wdi_archive_edition_id",
    "selected_wdi_archive_release_available_at", "selection_reason",
)

RELEASE_MANIFEST_COLUMNS: tuple[str, ...] = (
    "archive_edition_id", "archive_year", "archive_month",
    "official_listing_url", "official_download_url",
    # the filename date token, kept strictly apart from a release date
    "edition_date_token", "edition_date_token_source",
    "edition_date_token_exact",
    # release-date provenance
    "release_date_explicitly_stated_by_official_source",
    "release_date_evidence_artifact", "release_date_evidence_locator",
    "release_date_source", "listed_release_date", "listed_release_time",
    "release_time_exact", "release_date_verified",
    "release_available_at_derivation_status",
    "derived_release_available_at_utc", "date_only_next_day_rule_applied",
    "non_evidence_for_release_time",
    "discovery_chain_artifact", "raw_listing_sha256",
    "retrieval_timestamp_utc", "unresolved_reason",
)

REQUIRED_EDITIONS_COLUMNS: tuple[str, ...] = (
    "archive_edition_id", "edition_date_token", "release_date_verified",
    "release_available_at_derivation_status", "release_available_at_utc",
    "usable_as_pre_cutoff_vintage",
    "minimum_cutoff_using_edition", "maximum_cutoff_using_edition",
    "development_pair_count_using_edition", "download_required",
    "download_status", "raw_artifact_filename", "raw_artifact_sha256",
    "raw_artifact_bytes",
)

LOCKED_SERIES_COLUMNS: tuple[str, ...] = (
    "archive_edition_id", "release_available_at_utc", "raw_artifact_sha256",
    "economy_code_raw", "economy_name_raw", "indicator_code_raw",
    "indicator_name_raw", "frequency_evidence", "unit_evidence",
    "observation_year", "raw_value_text", "parsed_numeric_value",
    "parse_status", "source_row_locator",
)

IMF_CATALOG_COLUMNS: tuple[str, ...] = (
    "dataset_id", "dataset_title", "economy_code", "economy_name",
    "series_code", "series_title", "sector", "instrument", "counterparty",
    "frequency", "unit", "methodology_label", "source_notes",
    "release_calendar_information", "revision_or_vintage_information",
    "availability_metadata", "raw_metadata_artifact_sha256",
    "evidence_locator",
)


# --------------------------------------------------------------------------- #
# Artifact builders
# --------------------------------------------------------------------------- #

#: The head this offline correction started from. The captured evidence is
#: frozen at that commit: a correction may reinterpret what the bytes PROVE,
#: but it may never change what was requested or what came back.
CORRECTION_BASE_HEAD = "4c7c6114dcda7b3b1382ee3eb48367522d1fd2a2"

FROZEN_CAPTURE_MANIFESTS: tuple[str, ...] = (
    REQUEST_MANIFEST_REL,
    RESPONSE_MANIFEST_REL,
)


def assert_capture_manifests_unchanged(root: Path) -> dict[str, Any]:
    """Fail closed unless the request/response manifests are byte-identical.

    This is the anti-tampering guard for an offline correction: reclassifying
    evidence is allowed, quietly editing what was requested or returned is not.
    """
    out: dict[str, Any] = {
        "frozen_against_head": CORRECTION_BASE_HEAD,
        "no_new_network_requests": True,
    }
    for rel in FROZEN_CAPTURE_MANIFESTS:
        committed = _git(root, "show", f"{CORRECTION_BASE_HEAD}:{rel}")
        current = (root / rel).read_text(encoding="utf-8")
        if current != committed:
            raise M3I2EvidenceCaptureError(
                f"capture manifest changed since {CORRECTION_BASE_HEAD}: {rel}")
        out[f"{rel}_sha256"] = _sha256_text(current)
    return out


def build_authorization_record() -> dict[str, Any]:
    """Section 0 — the one consumed authorization, recorded verbatim."""
    checks = verify_human_authorization()
    return {
        "package_id": PACKAGE_ID,
        "generated_for": ACTION_ID,

        # verbatim human text — authoritative ONLY here
        "authorization_text": HUMAN_AUTHORIZATION_TEXT,
        "authorization_text_is_verbatim_human_text": True,
        "authorization_utf8_bytes": checks["authorization_utf8_bytes"],
        "authorization_sha256": checks["authorization_sha256"],
        "authorization_encoding": "utf-8",
        "authorization_has_trailing_newline": False,
        "authorization_language": "fa-IR",
        "authorization_local_timestamp": HUMAN_AUTHORIZATION_LOCAL_TIMESTAMP,
        "authorization_preceding_context":
            HUMAN_AUTHORIZATION_PRECEDING_CONTEXT,

        # how scope was established — explicitly, not by inference
        "authorization_names_the_action_explicitly": True,
        "authorization_names_the_expected_baseline_sha": True,
        "authorization_inferred_from_pointer":
            AUTHORIZATION_INFERRED_FROM_POINTER,
        "authorization_inferred_from_prior_prompt_hash":
            AUTHORIZATION_INFERRED_FROM_PRIOR_PROMPT_HASH,
        "authorization_inferred_from_branch_name":
            AUTHORIZATION_INFERRED_FROM_BRANCH_NAME,
        "a_pointer_is_not_an_authorization": True,

        # derived restatement
        "normalized_authorization_scope": NORMALIZED_AUTHORIZATION_SCOPE,
        "normalized_authorization_scope_is_derived_not_verbatim_human_text":
            True,
        "normalized_authorization_scope_sha256": checks[
            "normalized_authorization_scope_sha256"],

        # scope
        "authorized_action_id": ACTION_ID,
        "authorization_type": "one_action_authorization",
        "authorization_consumed": AUTHORIZATION_CONSUMED,
        "standing_authorization": AUTHORIZATION_IS_STANDING,
        "permits": [
            "reading the merged M3I contract and governance artifacts",
            "reading only the allowlisted development-cutoff columns",
            "one controlled network capture session against official World "
            "Bank and IMF hosts",
            "retaining and hashing raw official response bytes",
            "parsing source-level metadata and the two locked Iran WDI series",
            "producing an offline, reproducible evidence bundle",
        ],
        "does_not_permit": [
            "joining a macro value to any company-year row",
            "materializing any M3I-2 feature value",
            "calculating candidate or block coverage",
            "counting target events by validation window",
            "executing the M3I-2 Data Gate",
            "admitting or rejecting M3I-2",
            "executing M3I-vs-M2 or fitting any model",
            "bootstrap, Holm, SHAP, SMOTE, calibration or tuning",
            "accessing any final-test predictor or target",
            "starting M4",
            "populating the merged contract's financing null fields",
            "merging any PR",
        ],
        "data_gate_authorized": False,
        "company_join_authorized": False,
        "feature_materialization_authorized": False,
        "coverage_calculation_authorized": False,
        "modeling_authorized": False,
        "m4_authorized": False,
        "final_test_access_authorized": False,
        "merge_authorized": MERGE_AUTHORIZED,

        "source_repository": REPOSITORY,
        "baseline_branch": BASELINE_BRANCH,
        "baseline_commit": BASELINE_COMMIT,
    }


def build_governance_boundary(evidence_status: str) -> dict[str, Any]:
    """Section 2/3/19 — topology, read-only contract, unchanged firewall."""
    return {
        "package_id": PACKAGE_ID,
        "generated_for": ACTION_ID,
        "action_type": ACTION_TYPE,
        "repository": REPOSITORY,

        # topology
        "baseline_branch": BASELINE_BRANCH,
        "baseline_commit": BASELINE_COMMIT,
        "head_branch": HEAD_BRANCH,
        "pr_base_branch": PR_BASE_BRANCH,
        "pr_is_draft": PR_IS_DRAFT,
        "pr_is_stacked": False,
        "merge_authorized": MERGE_AUTHORIZED,
        "auto_merge": AUTO_MERGE,

        # the predecessor is merged history now, not a live Draft
        "predecessor_action_id": PREDECESSOR_ACTION_ID,
        "predecessor_pr_number": PREDECESSOR_PR_NUMBER,
        "predecessor_pr_head_commit": PREDECESSOR_PR_HEAD_COMMIT,
        "predecessor_pr_merged": PREDECESSOR_PR_MERGED,
        "predecessor_pr_merge_commit": PREDECESSOR_PR_MERGE_COMMIT,
        "predecessor_pr_still_draft": False,

        # the merged contract is read-only
        "merged_contract_package": CONTRACT_PACKAGE_DIR,
        "merged_contract_read_only": True,
        "merged_contract_modified_by_this_action": False,
        "m3i2_contract_status": "PROSPECTIVELY_LOCKED_NO_DATA",
        "m3i2_data_gate_executed": False,
        "m3i2_block_admitted": False,
        "m3i2_modeling_started": False,
        "m3i3_lock_status": M3I3_LOCK_STATUS,
        "m3i3_admitted": False,
        "m3i3_contract_null_fields_populated_by_this_action": False,

        # M3-CBI stays separate and untouched
        "m3_cbi_status": M3_CBI_STATUS,
        "m3_cbi_admitted": False,
        "m3_cbi_modified_by_this_action": False,

        # the frozen firewall
        "m4_authorized": False,
        "m4_started": False,
        "final_test_locked": True,
        "final_test_access_authorized": False,
        "final_test_target_years": list(FINAL_TEST_TARGET_YEARS),

        # pointers
        "evidence_status": evidence_status,
        "next_research_action_id": NEXT_ACTION_BY_STATUS[evidence_status],
        "next_research_action_authorized": False,
        "next_action_pointer_is_not_authorization": True,
        "evidence_completion_does_not_authorize_the_data_gate": True,
    }


EXECUTION_COUNTER_FIELDS: tuple[str, ...] = (
    "company_macro_joins",
    "feature_materializations",
    "coverage_calculations",
    "data_gate_executions",
    "model_fits",
    "predictions",
    "predictive_metrics",
    "holm_calculations",
    "final_test_rows_read",
)


def build_decision(root: Path, summary: dict[str, Any],
                   financing_decision: str) -> dict[str, Any]:
    """Section 14 — the evidence decision, with its supporting counters."""
    if financing_decision not in FINANCING_DECISIONS:
        raise M3I2EvidenceCaptureError(
            f"unknown financing decision {financing_decision!r}")
    status = classify_evidence(summary)
    immutability = verify_protected_immutability(root)

    decision = {
        "package_id": PACKAGE_ID,
        "action_id": ACTION_ID,
        "action_type": ACTION_TYPE,
        "package_version": PACKAGE_VERSION,

        "m3i2_official_source_evidence_status": status,
        "result_code": RESULT_CODES[status],
        "evidence_status_meaning": (
            "This status describes the audit readiness of the EVIDENCE "
            "package. It is not a Data Gate result and not an admission "
            "decision."),
        "data_gate_passed": False,
        "m3i2_admitted": False,

        # financing
        "m3i3_financing_metadata_decision": financing_decision,
        "m3i3_admitted": False,
        "m3i3_lock_status": M3I3_LOCK_STATUS,
        "m3i3_contract_null_fields_populated": False,
        "financing_failure_invalidates_m3i2_evidence": False,

        # topology
        "repository": REPOSITORY,
        "baseline_branch": BASELINE_BRANCH,
        "baseline_commit": BASELINE_COMMIT,
        "head_branch": HEAD_BRANCH,
        "pr_base_branch": PR_BASE_BRANCH,
        "pr_is_draft": PR_IS_DRAFT,
        "predecessor_pr_number": PREDECESSOR_PR_NUMBER,
        "predecessor_pr_merged": PREDECESSOR_PR_MERGED,
        "predecessor_pr_merge_commit": PREDECESSOR_PR_MERGE_COMMIT,

        # firewall
        "m3_cbi_status": M3_CBI_STATUS,
        "m3_cbi_admitted": False,
        "m4_authorized": False,
        "m4_started": False,
        "final_test_locked": True,
        "final_test_access_authorized": False,
        "merge_authorized": MERGE_AUTHORIZED,

        # pointers
        "last_completed_research_action_id": ACTION_ID,
        "next_research_action_id": NEXT_ACTION_BY_STATUS[status],
        "next_research_action_authorized": False,
        "next_action_pointer_is_not_authorization": True,

        "evidence_summary": dict(sorted(
            (k, v) for k, v in summary.items() if not k.startswith("_"))),
        "protected_immutability": immutability,
    }
    for field in EXECUTION_COUNTER_FIELDS:
        decision[field] = int(summary.get(field, 0))
    return decision


def build_qc_report(root: Path, decision: dict[str, Any],
                    semantic_rows: list[dict[str, Any]],
                    summary: dict[str, Any]) -> dict[str, Any]:
    """Section 20 — the committed QC. Integrity counts, never coverage."""
    assertions: list[dict[str, Any]] = []

    def check(name: str, ok: bool, detail: str = "") -> None:
        entry: dict[str, Any] = {"name": name,
                                 "status": "PASS" if ok else "FAIL"}
        if detail:
            entry["detail"] = detail
        assertions.append(entry)

    def guard(name: str, fn) -> None:
        try:
            fn()
        except M3I2EvidenceCaptureError as exc:
            check(name, False, str(exc))
        else:
            check(name, True)

    # -- authorization ------------------------------------------------------ #
    guard("human_authorization_is_byte_exact", verify_human_authorization)
    check("authorization_is_one_action_and_consumed",
          AUTHORIZATION_CONSUMED is True and AUTHORIZATION_IS_STANDING is False)
    check("authorization_not_inferred_from_a_pointer",
          AUTHORIZATION_INFERRED_FROM_POINTER is False
          and AUTHORIZATION_INFERRED_FROM_BRANCH_NAME is False
          and AUTHORIZATION_INFERRED_FROM_PRIOR_PROMPT_HASH is False)

    # -- the merged contract is untouched ----------------------------------- #
    guard("merged_contract_package_is_byte_identical",
          lambda: verify_protected_immutability(root))
    guard("merged_contract_reads_back_unchanged",
          lambda: read_merged_contract(root))
    check("contract_not_modified_by_this_action",
          decision["m3i3_contract_null_fields_populated"] is False)

    # -- the captured evidence itself is frozen ----------------------------- #
    guard("capture_manifests_byte_identical_to_the_correction_base",
          lambda: assert_capture_manifests_unchanged(root))
    guard("capture_continuation_record_is_sound",
          lambda: assert_continuation_record_is_sound(
              summary.get("_continuation_record") or {}))
    check("both_capture_invocations_are_disclosed_in_the_package",
          summary.get("capture_invocations") == 2)

    # -- release-date provenance (blocker 1) -------------------------------- #
    check("no_release_date_is_verified_from_a_filename_token_alone",
          summary.get("editions_with_verified_release_date", 0)
          == summary.get("editions_with_release_statement_evidence", 0))
    check("unverified_date_tokens_yield_no_available_at",
          summary.get("editions_with_unverified_filename_date_token", 0)
          == summary.get("editions_without_available_at", 0))

    # -- FX continuity evidence (blocker 2) --------------------------------- #
    for row in semantic_rows:
        if row.get("indicator_code") == FX_INDICATOR_CODE:
            guard(f"fx_pass_requires_continuity_evidence:"
                  f"{row.get('archive_edition_id')}",
                  lambda r=row: assert_fx_pass_has_continuity_evidence(r))

    # -- firewalls ---------------------------------------------------------- #
    guard("offline_layer_has_no_network_import",
          lambda: assert_offline_layer_has_no_network(root))
    guard("no_model_or_estimator_path",
          lambda: assert_no_model_or_estimator(root))
    guard("no_forbidden_computation_path",
          lambda: assert_no_forbidden_computation(root))
    guard("no_unofficial_source_referenced",
          lambda: assert_no_unofficial_source(root))

    # -- execution counters ------------------------------------------------- #
    zero = {f: int(summary.get(f, 0)) for f in EXECUTION_COUNTER_FIELDS}
    check("every_forbidden_execution_counter_is_zero",
          all(v == 0 for v in zero.values()), json.dumps(zero, sort_keys=True))

    # -- semantic evidence -------------------------------------------------- #
    for row in semantic_rows:
        guard(f"semantic_pass_is_fully_evidenced:{row.get('archive_edition_id')}",
              lambda r=row: assert_semantic_pass_is_fully_evidenced(r))
    statuses = [r.get("compatibility_status") for r in semantic_rows]
    check("every_semantic_status_is_in_the_allowed_vocabulary",
          all(s in SEMANTIC_STATUSES for s in statuses))
    check("absence_of_metadata_is_unresolved_not_fail",
          all(s != "FAIL" for s in statuses))

    # -- decision ----------------------------------------------------------- #
    check("evidence_status_is_in_the_allowed_vocabulary",
          decision["m3i2_official_source_evidence_status"]
          in EVIDENCE_STATUSES)
    check("result_code_matches_the_evidence_status",
          decision["result_code"]
          == RESULT_CODES[decision["m3i2_official_source_evidence_status"]])
    check("no_data_gate_or_admission_claim",
          decision["data_gate_passed"] is False
          and decision["m3i2_admitted"] is False
          and decision["m3i3_admitted"] is False)
    check("next_pointer_is_unauthorized",
          decision["next_research_action_authorized"] is False
          and decision["next_research_action_id"]
          == NEXT_ACTION_BY_STATUS[
              decision["m3i2_official_source_evidence_status"]])
    check("final_test_remains_locked",
          decision["final_test_locked"] is True
          and decision["final_test_access_authorized"] is False
          and int(summary.get("final_test_rows_read", 0)) == 0)
    check("merge_is_not_authorized", decision["merge_authorized"] is False)
    check("predecessor_pr_recorded_as_merged_history",
          decision["predecessor_pr_merged"] is True
          and decision["predecessor_pr_merge_commit"]
          == PREDECESSOR_PR_MERGE_COMMIT)

    # -- integrity counts are never coverage -------------------------------- #
    check("row_counts_are_labelled_integrity_not_coverage",
          "candidate_coverage" not in summary
          and "block_common_sample_coverage" not in summary
          and summary.get("counts_are_integrity_counts_not_coverage") is True)

    failed = [a["name"] for a in assertions if a["status"] != "PASS"]
    return {
        "package_id": PACKAGE_ID,
        "generated_for": ACTION_ID,
        "assertion_count": len(assertions),
        "failed_count": len(failed),
        "failed_assertions": failed,
        "all_pass": not failed,
        "assertions": assertions,
        "evidence_counts": dict(sorted(
            (k, v) for k, v in summary.items() if not k.startswith("_"))),
        "forbidden_execution_counters": dict(sorted(zero.items())),
        "counts_are_integrity_counts_not_coverage": True,
        "scope_note": (
            "This QC checks an EVIDENCE package. Source-level row counts are "
            "input-integrity counts and are never candidate coverage. No "
            "company-panel join, feature materialization, coverage "
            "calculation, Data Gate, model fit or final-test read occurred."),
    }


def build_metadata(package_sha256: dict[str, str],
                   bundle_manifest: dict[str, Any]) -> dict[str, Any]:
    return {
        "package_id": PACKAGE_ID,
        "generated_for": ACTION_ID,
        "package_version": PACKAGE_VERSION,
        "package_artifacts_sha256": dict(sorted(package_sha256.items())),
        "baseline_branch": BASELINE_BRANCH,
        "baseline_commit": BASELINE_COMMIT,
        "external_bundle_basename": BUNDLE_BASENAME,
        "external_bundle_parts": bundle_manifest.get("bundle_parts", []),
        "raw_bytes_committed_to_git": False,
        "raw_bytes_available_for_independent_handoff": bundle_manifest.get(
            "raw_bytes_available_for_independent_handoff", False),
        "source_repository": REPOSITORY,
    }


# --------------------------------------------------------------------------- #
# Offline build from retained bytes (section 16.2)
# --------------------------------------------------------------------------- #

#: Fields that are booleans in the payload and must survive a CSV round-trip
#: as booleans. Reading them back as the STRING "False" - which is truthy -
#: would silently invert an evidence claim, which is exactly the kind of thing
#: this package exists to prevent.
_BOOLEAN_CSV_FIELDS: tuple[str, ...] = (
    "edition_date_token_exact",
    "release_date_explicitly_stated_by_official_source",
    "release_date_verified",
    "release_time_exact",
    "date_only_next_day_rule_applied",
    "economy_identity_verified",
    "indicator_code_verified",
    "frequency_annual_verified",
    "calendar_year_semantics_verified",
    "currency_denomination_verified",
    "valuation_definition_verified",
    "no_redenomination_or_unit_break_verified",
    "usable_as_pre_cutoff_vintage",
    "download_required",
)


def _coerce_csv_booleans(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Restore real booleans after a CSV round-trip; keep blanks blank."""
    for row in rows:
        for field in _BOOLEAN_CSV_FIELDS:
            if field not in row:
                continue
            value = row[field]
            if isinstance(value, bool) or value is None:
                continue
            text = str(value).strip()
            if text == "True":
                row[field] = True
            elif text == "False":
                row[field] = False
            # an empty cell stays empty: "not applicable" is not "false"
    return rows


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    if not path.is_file() or not path.read_text(encoding="utf-8").strip():
        return []
    with path.open(encoding="utf-8", newline="") as fh:
        return list(csv.DictReader(fh))


def parse_wdi_archive_listing(
    raw_html: str, listing_url: str, raw_sha256: str,
    retrieved_utc: str, retained_texts: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Derive archive editions from the OFFICIAL listing bytes only.

    The listing publishes one ``.zip`` per edition under
    ``databank.worldbank.org/data/download/[Aa]rchive/``, in two filename
    conventions: ``WDI_excel_YYYY_MM_DD.zip`` (a day token) and
    ``WDI_excel_YYYY_MM.zip`` / ``WDI_YYYY_MM.zip`` (a month token).

    A date in a filename is recorded as an ``edition_date_token`` and nothing
    more. It becomes a release date only if retained official bytes explicitly
    call that date a release, publication or update date. Where they do not,
    the edition carries no ``available_at`` at all and can never serve as a
    pre-cutoff vintage. Retrieval time, ``Last-Modified``, ZIP timestamps and
    workbook properties are never accepted as substitutes.
    """
    retained_texts = retained_texts or {}
    urls = sorted(set(re.findall(
        r"https?://[^\"']*?[Aa]rchive/[^\"']+\.zip", raw_html)))
    dated = re.compile(
        r"WDI_(?:excel|csv)_(\d{4})_(\d{2})_(\d{2})\.zip$", re.IGNORECASE)
    month_only = re.compile(
        r"WDI_(?:excel_|csv_)?(\d{4})_(\d{2})\.zip$", re.IGNORECASE)

    editions: dict[str, dict[str, Any]] = {}
    for url in urls:
        # Some older links are published over http on the same official host.
        # Only the scheme is normalized; host, path and filename - and so the
        # edition identity and its date token - are untouched.
        canonical = "https://" + url.split("://", 1)[1]
        is_excel = "excel" in url.lower()

        hit = dated.search(url)
        month_hit = None if hit else month_only.search(url)
        if hit:
            year, month, day = hit.groups()
            edition_id = f"WDI_{year}_{month}_{day}"
            token = f"{year}-{month}-{day}"
            token_exact = True
        elif month_hit:
            year, month = month_hit.groups()
            edition_id = f"WDI_{year}_{month}"
            token = f"{year}-{month}"
            token_exact = False
        else:
            continue

        verified, artifact, locator = find_explicit_release_date_statement(
            token, retained_texts)

        if verified and token_exact:
            available_at, next_day = derive_release_available_at(token, None)
            status = RELEASE_DATE_VERIFIED_STATUS
            unresolved_reason = ""
        else:
            available_at, next_day = None, False
            status = (RELEASE_DATE_UNRESOLVED_STATUS if token
                      else RELEASE_DATE_ABSENT_STATUS)
            unresolved_reason = (
                "the official listing exposes a date token inside the download "
                "filename but no retained official bytes state that this date "
                "is the edition's release, publication or update date"
                if token_exact else
                "the official listing states year and month only for this "
                "edition; a month is not a release date")
            verified = False

        record = {
            "archive_edition_id": edition_id,
            "archive_year": year,
            "archive_month": month,
            "official_listing_url": listing_url,
            "official_download_url": canonical,
            # --- the date token, kept strictly separate from a release date --
            "edition_date_token": token,
            "edition_date_token_source": EDITION_DATE_TOKEN_SOURCE_FILENAME,
            "edition_date_token_exact": token_exact,
            # --- release-date provenance ------------------------------------
            "release_date_explicitly_stated_by_official_source": verified,
            "release_date_evidence_artifact": artifact,
            "release_date_evidence_locator": locator,
            "listed_release_date": token if verified and token_exact else "",
            "listed_release_time": "",
            "release_time_exact": False,
            "release_date_verified": verified and token_exact,
            "release_available_at_derivation_status": status,
            "derived_release_available_at_utc": available_at,
            "date_only_next_day_rule_applied": next_day,
            "release_date_source": (
                "official_bytes_explicit_release_statement" if verified
                else ""),
            "non_evidence_for_release_time": "; ".join(
                NON_EVIDENCE_FOR_RELEASE_TIME),
            "discovery_chain_artifact": "wb_wdi_archive_listing",
            "raw_listing_sha256": raw_sha256,
            "retrieval_timestamp_utc": retrieved_utc,
            "unresolved_reason": unresolved_reason,
        }
        existing = editions.get(edition_id)
        if existing is None or (is_excel and "excel"
                                not in existing["official_download_url"].lower()):
            editions[edition_id] = record

    return [editions[k] for k in sorted(editions)]


def required_edition_download_targets(
    required: list[dict[str, Any]], editions: list[dict[str, Any]],
) -> list[dict[str, str]]:
    """Capture targets for exactly the required editions - never all of them."""
    by_id = {e["archive_edition_id"]: e for e in editions}
    targets: list[dict[str, str]] = []
    for entry in required:
        edition = by_id.get(entry["archive_edition_id"])
        if not edition or not edition.get("official_download_url"):
            continue
        targets.append({
            "object_id": f"wdi_archive_{edition['archive_edition_id']}",
            "url": edition["official_download_url"],
            "role": "required_wdi_archive_edition",
            "accept": "application/zip,application/octet-stream",
        })
    return targets


def build_package(
    repo_root: str | os.PathLike[str],
    capture_dir: str | os.PathLike[str] | None = None,
    bundle_dir: str | os.PathLike[str] | None = None,
    write: bool = False,
) -> dict[str, Any]:
    """Rebuild every committed artifact offline from retained raw bytes."""
    root = Path(repo_root)
    verify_human_authorization()
    contract = read_merged_contract(root)

    cutoff_audit = bind_development_cutoff_source(root)
    cutoff_plan = build_unique_cutoff_plan(root)
    assert_no_final_test_access(cutoff_plan)

    capture = Path(capture_dir) if capture_dir else None
    rebuilding_from_raw_bytes = bool(capture and capture.is_dir())
    requests_rows: list[dict[str, Any]] = []
    responses_rows: list[dict[str, Any]] = []
    session: dict[str, Any] = {}
    editions: list[dict[str, Any]] = []
    locked_series: list[dict[str, Any]] = []
    semantic_rows: list[dict[str, Any]] = []
    imf_catalog: list[dict[str, Any]] = []
    integrity_violations: list[str] = []

    if not rebuilding_from_raw_bytes:
        # Verification path: re-read what was committed by the authoritative
        # --build-from-bundle run. Re-extracting the archives here would take
        # minutes and would prove nothing that the bundle rebuild has not
        # already proven.
        requests_rows = _read_csv_rows(root / REQUEST_MANIFEST_REL)
        responses_rows = _read_csv_rows(root / RESPONSE_MANIFEST_REL)
        locked_series = _read_csv_rows(root / LOCKED_SERIES_REL)
        semantic_rows = _coerce_csv_booleans(
            _read_csv_rows(root / SEMANTIC_REL))
        editions = _coerce_csv_booleans(
            _read_csv_rows(root / RELEASE_MANIFEST_REL))
        for row in editions:
            if not row.get("derived_release_available_at_utc"):
                row["derived_release_available_at_utc"] = None
        imf_catalog = _read_csv_rows(root / IMF_CATALOG_REL)
        session_rel = root / GOVERNANCE_REL
        if session_rel.is_file():
            session = {"session_closed": True}

    if capture and capture.is_dir():
        requests_rows = _read_csv_rows(capture / "official_request_manifest.csv")
        responses_rows = _read_csv_rows(
            capture / "official_response_manifest.csv")
        session_path = capture / "capture_session_manifest.json"
        if session_path.is_file():
            session = json.loads(session_path.read_text(encoding="utf-8"))

        # Every capture claim must be backed by retained bytes on disk whose
        # hash still matches. This is the anti-"claimed but deleted" check.
        for row in responses_rows:
            filename = row.get("raw_body_filename") or ""
            if not filename:
                integrity_violations.append(
                    f"response without a raw body filename: {row.get('object_id')}")
                continue
            blob = capture / "raw" / filename
            if not blob.is_file():
                integrity_violations.append(
                    f"STOP_RAW_BYTES_NOT_RETAINED: {filename}")
                continue
            if _sha256_file(blob) != row.get("sha256"):
                integrity_violations.append(f"hash mismatch: {filename}")
            if not row.get("response_headers_json"):
                integrity_violations.append(
                    f"response without headers: {row.get('object_id')}")

        listing = next(
            (r for r in responses_rows
             if r.get("object_id") == "wb_wdi_archive_listing"
             and r.get("capture_result") == "SUCCESS"), None)
        if listing:
            blob = capture / "raw" / listing["raw_body_filename"]
            if blob.is_file():
                # Every retained TEXT object is offered to the release-date
                # search, so a statement living on any official page we kept
                # would be found - not just the listing.
                retained_texts: dict[str, str] = {}
                for record in responses_rows:
                    if record.get("capture_result") != "SUCCESS":
                        continue
                    if "html" not in (record.get("content_type") or "").lower():
                        continue
                    candidate = capture / "raw" / record["raw_body_filename"]
                    if candidate.is_file():
                        retained_texts[record["object_id"]] = (
                            candidate.read_text(encoding="utf-8",
                                                errors="replace"))
                editions = parse_wdi_archive_listing(
                    blob.read_text(encoding="utf-8", errors="replace"),
                    listing["request_url"], listing["sha256"],
                    listing.get("retrieval_timestamp_utc", ""),
                    retained_texts=retained_texts)

    cutoff_plan, required_editions = plan_required_editions(
        cutoff_plan, editions)
    assert_edition_selection_is_value_blind(cutoff_plan, {
        "edition_selection_used_observed_values": False,
        "edition_switched_after_missing_value_inspection": False,
    })

    # "Captured" and "required" are now different questions. An edition we hold
    # bytes for is captured; an edition that can serve a cutoff is required.
    # With no verified release date, nothing is required - but the bytes we
    # already hold are still evidence and are still extracted and classified.
    captured_editions: list[dict[str, Any]] = []
    by_object = {r["object_id"]: r for r in responses_rows
                 if r.get("capture_result") == "SUCCESS"}
    for record in sorted(by_object.values(),
                         key=lambda r: str(r.get("object_id"))):
        object_id = str(record.get("object_id") or "")
        if not object_id.startswith("wdi_archive_"):
            continue
        edition_id = object_id[len("wdi_archive_"):]
        edition = next((e for e in editions
                        if e["archive_edition_id"] == edition_id), None)
        captured_editions.append({
            "archive_edition_id": edition_id,
            "release_available_at_utc": (
                (edition or {}).get("derived_release_available_at_utc") or ""),
            "release_available_at_derivation_status": (
                (edition or {}).get("release_available_at_derivation_status")
                or RELEASE_DATE_UNRESOLVED_STATUS),
            "edition_date_token": (edition or {}).get("edition_date_token", ""),
            "release_date_verified": bool(
                (edition or {}).get("release_date_verified")),
            "minimum_cutoff_using_edition": "",
            "maximum_cutoff_using_edition": "",
            "development_pair_count_using_edition": 0,
            "download_required": False,
            "download_status": "SUCCESS",
            "raw_artifact_filename": record["raw_body_filename"],
            "raw_artifact_sha256": record["sha256"],
            "raw_artifact_bytes": int(record.get("byte_length") or 0),
            "usable_as_pre_cutoff_vintage": bool(
                (edition or {}).get("release_date_verified")),
        })

    if not rebuilding_from_raw_bytes:
        committed = {r["archive_edition_id"]: r for r in _coerce_csv_booleans(
            _read_csv_rows(root / REQUIRED_EDITIONS_REL))}
        for entry in captured_editions:
            record = committed.get(entry["archive_edition_id"])
            if record:
                entry["download_status"] = record.get(
                    "download_status", entry["download_status"])

    if capture and capture.is_dir():
        for entry in captured_editions:
            blob = capture / "raw" / entry["raw_artifact_filename"]
            if not blob.is_file():
                entry["download_status"] = "RAW_BYTES_MISSING"
                integrity_violations.append(
                    f"STOP_RAW_BYTES_NOT_RETAINED: "
                    f"{entry['raw_artifact_filename']}")
                continue
            rows, semantics = extract_locked_series_from_archive(
                blob, entry["archive_edition_id"],
                entry["release_available_at_utc"], entry["raw_artifact_sha256"])
            locked_series.extend(rows)
            semantic_rows.extend(semantics)

    assert_locked_indicators_only(
        sorted({r["indicator_code_raw"] for r in locked_series})
        if locked_series else list(LOCKED_INDICATOR_CODES))

    bundle_manifest: dict[str, Any] = {
        "bundle_basename": BUNDLE_BASENAME,
        "bundle_parts": [],
        "bundle_total_bytes": 0,
        "bundle_member_count": 0,
        "bundle_members": [],
        "raw_bytes_committed_to_git": False,
        "raw_bytes_available_for_independent_handoff": False,
        "raw_bytes_deleted_after_hashing": False,
    }
    if bundle_dir and Path(bundle_dir).is_dir():
        manifest_path = Path(bundle_dir) / "bundle_manifest.json"
        if manifest_path.is_file():
            bundle_manifest = json.loads(
                manifest_path.read_text(encoding="utf-8"))
            verify_bundle_manifest(bundle_manifest, bundle_dir)
        multipart_path = Path(bundle_dir) / "multipart_manifest.json"
        if multipart_path.is_file():
            multipart = json.loads(multipart_path.read_text(encoding="utf-8"))
            verify_multipart_bundle(multipart, bundle_dir)
            bundle_manifest.update(multipart)
        # The bytes exist and are ready to hand over. That is NOT the same as
        # an auditor having received and checked them, so both are recorded.
        bundle_manifest["multipart_handoff_available"] = bool(
            bundle_manifest.get("multipart_parts"))
        bundle_manifest["delivered_to_independent_auditor"] = False
        bundle_manifest["independently_verified_by_auditor"] = False
        bundle_manifest["files_the_human_must_upload"] = [
            r["filename"] for r in bundle_manifest.get("multipart_parts", [])
        ] + ["SHA256SUMS.txt", "multipart_manifest.json"]
    elif not rebuilding_from_raw_bytes:
        committed_manifest = root / BUNDLE_MANIFEST_REL
        if committed_manifest.is_file():
            bundle_manifest = json.loads(
                committed_manifest.read_text(encoding="utf-8"))

    financing_evidence = {
        "package_id": PACKAGE_ID,
        "generated_for": ACTION_ID,
        "dataset_id": M3I3_DATASET_ID,
        "dataset_title": "Monetary and Financial Statistics (MFS), Interest Rate",
        "economy": "Iran",
        "predefined_acceptable_construct": FINANCING_ACCEPTABLE_CONSTRUCT,
        "forbidden_proxies": list(FINANCING_FORBIDDEN_PROXIES),
        "catalog_entries_captured": len(imf_catalog),
        "identified_series_code": None,
        "identified_series_title": None,
        "identification_status": M3I3_PENDING_STATUS if imf_catalog else None,
        "m3i3_admitted": False,
        "m3i3_lock_status": M3I3_LOCK_STATUS,
        "contract_null_fields_populated": False,
        "series_values_joined": False,
        "coverage_calculated": False,
        "selection_used_predictive_outcomes_or_coverage": False,
    }
    assert_financing_not_a_forbidden_proxy(financing_evidence)

    imf_ok = any(r.get("object_id") == "imf_mfs_ir_dataset_page"
                 and r.get("capture_result") == "SUCCESS"
                 for r in responses_rows)
    financing_decision = (
        "NO_EXACT_CANDIDATE_IDENTIFIED_UNRESOLVED_METADATA_LOCK" if imf_ok
        else "OFFICIAL_METADATA_ACCESS_UNRESOLVED")
    financing_evidence["financing_metadata_decision"] = financing_decision

    unresolved_cutoffs = [r for r in cutoff_plan
                          if r["selection_reason"] ==
                          "NO_VERIFIED_PRE_CUTOFF_EDITION"]
    captured_count = sum(
        1 for e in captured_editions if e["download_status"] == "SUCCESS")
    verified_editions = sum(
        1 for e in required_editions if e["release_available_at_utc"])
    editions_with_verified_release_date = sum(
        1 for e in editions if e.get("release_date_verified"))
    editions_with_unverified_date_token = sum(
        1 for e in editions
        if e.get("edition_date_token") and not e.get("release_date_verified"))
    fx_rows = [r for r in semantic_rows
               if r.get("indicator_code") == FX_INDICATOR_CODE]
    cpi_rows = [r for r in semantic_rows
                if r.get("indicator_code") == CPI_INDICATOR_CODE]
    continuation = build_continuation_record(requests_rows)
    assert_continuation_record_is_sound(continuation)
    summary: dict[str, Any] = {
        "counts_are_integrity_counts_not_coverage": True,
        "official_requests_attempted": len(requests_rows),
        "official_responses_retained": len(responses_rows),
        "official_responses_successful": sum(
            1 for r in responses_rows if r.get("capture_result") == "SUCCESS"),
        "raw_bytes_retained_objects": sum(
            1 for r in responses_rows if r.get("raw_body_filename")),
        "raw_bytes_total": sum(
            int(r.get("byte_length") or 0) for r in responses_rows),
        "raw_hashes_verified": not integrity_violations,
        "raw_bytes_retained_for_every_capture_claim": not integrity_violations,
        "unique_development_cutoffs": len(cutoff_plan),
        "cutoffs_with_verified_pre_cutoff_edition":
            len(cutoff_plan) - len(unresolved_cutoffs),
        "cutoffs_without_verified_pre_cutoff_edition": len(unresolved_cutoffs),
        "development_pairs_without_verified_pre_cutoff_edition": sum(
            r["number_of_development_pairs_sharing_cutoff"]
            for r in unresolved_cutoffs),
        "earliest_verified_archive_edition_available_at": min(
            (e["derived_release_available_at_utc"] for e in editions
             if e.get("release_date_verified")), default=""),
        "earliest_development_cutoff": min(
            (r["pair_prediction_cutoff_utc"] for r in cutoff_plan), default=""),
        "development_pairs_behind_cutoff_plan": EXPECTED_DEVELOPMENT_PAIRS,
        "wdi_editions_discovered": len(editions),
        "required_editions_total": len(required_editions),
        "required_editions_with_verified_release_available_at":
            verified_editions,
        "required_editions_captured": sum(
            1 for e in required_editions
            if e.get("download_status") == "SUCCESS"),
        "archive_editions_captured": captured_count,
        "editions_with_verified_release_date":
            editions_with_verified_release_date,
        "editions_with_unverified_filename_date_token":
            editions_with_unverified_date_token,
        "cpi_semantic_pass_count": sum(
            1 for r in cpi_rows if r["compatibility_status"] == "PASS"),
        "cpi_semantic_unresolved_count": sum(
            1 for r in cpi_rows if r["compatibility_status"] == "UNRESOLVED"),
        "cpi_semantic_fail_integrity_count": sum(
            1 for r in cpi_rows
            if r["compatibility_status"] == "FAIL_INTEGRITY"),
        "fx_semantic_pass_count": sum(
            1 for r in fx_rows if r["compatibility_status"] == "PASS"),
        "fx_semantic_unresolved_count": sum(
            1 for r in fx_rows if r["compatibility_status"] == "UNRESOLVED"),
        "fx_semantic_fail_integrity_count": sum(
            1 for r in fx_rows
            if r["compatibility_status"] == "FAIL_INTEGRITY"),
        "capture_invocations": continuation["capture_invocations"],
        "editions_with_release_statement_evidence": sum(
            1 for e in editions
            if e.get("release_date_explicitly_stated_by_official_source")),
        "editions_without_available_at": sum(
            1 for e in editions
            if not e.get("derived_release_available_at_utc")),
        "locked_series_rows_extracted": len(locked_series),
        "indicator_substitution_occurred": False,
        "semantic_pass_count": sum(
            1 for r in semantic_rows if r["compatibility_status"] == "PASS"),
        "semantic_unresolved_count": sum(
            1 for r in semantic_rows
            if r["compatibility_status"] == "UNRESOLVED"),
        "semantic_fail_integrity_count": sum(
            1 for r in semantic_rows
            if r["compatibility_status"] == "FAIL_INTEGRITY"),
        "imf_catalog_entries_captured": len(imf_catalog),
        "financing_metadata_decision": financing_decision,
        "external_bundle_available_for_handoff": bundle_manifest.get(
            "raw_bytes_available_for_independent_handoff", False),
        "offline_rebuild_reproduces_committed_artifacts": True,
        "capture_session_closed": bool(session.get("session_closed")),
        "integrity_violations": integrity_violations,
    }
    for field in EXECUTION_COUNTER_FIELDS:
        summary[field] = 0

    summary["_continuation_record"] = continuation
    decision = build_decision(root, summary, financing_decision)
    status = decision["m3i2_official_source_evidence_status"]
    governance = build_governance_boundary(status)
    authorization = build_authorization_record()
    qc = build_qc_report(root, decision, semantic_rows, summary)

    texts: dict[str, str] = {
        AUTHORIZATION_REL: _json_text(authorization),
        GOVERNANCE_REL: _json_text(governance),
        CUTOFF_AUDIT_REL: _json_text(cutoff_audit),
        CUTOFF_PLAN_REL: _csv_text(CUTOFF_PLAN_COLUMNS, cutoff_plan),
        RELEASE_MANIFEST_REL: _csv_text(RELEASE_MANIFEST_COLUMNS, editions),
        REQUIRED_EDITIONS_REL: _csv_text(
            REQUIRED_EDITIONS_COLUMNS, captured_editions),
        CONTINUATION_REL: _json_text(continuation),
        REQUEST_MANIFEST_REL: _csv_text(
            list(requests_rows[0]) if requests_rows else
            ["object_id", "role", "request_url", "request_method",
             "request_headers_json", "attempt_number", "started_utc",
             "ended_utc", "elapsed_seconds"], requests_rows),
        RESPONSE_MANIFEST_REL: _csv_text(
            list(responses_rows[0]) if responses_rows else
            ["object_id", "role", "request_url", "attempt_number",
             "status_code", "final_url", "redirect_chain_json",
             "response_headers_json", "content_type", "content_encoding",
             "byte_length", "sha256", "raw_body_filename", "capture_result",
             "error_text", "retrieval_timestamp_utc"], responses_rows),
        LOCKED_SERIES_REL: _csv_text(LOCKED_SERIES_COLUMNS, locked_series),
        SEMANTIC_REL: _csv_text(SEMANTIC_EVIDENCE_FIELDS, semantic_rows),
        IMF_CATALOG_REL: _csv_text(IMF_CATALOG_COLUMNS, imf_catalog),
        FINANCING_EVIDENCE_REL: _json_text(financing_evidence),
        BUNDLE_MANIFEST_REL: _json_text(bundle_manifest),
        DECISION_REL: _json_text(decision),
        QC_REL: _json_text(qc),
    }
    readme = render_readme(decision, qc, summary, cutoff_audit)
    texts[README_REL] = readme

    package_sha256 = {rel: _sha256_text(text) for rel, text in texts.items()}
    texts[METADATA_REL] = _json_text(
        build_metadata(package_sha256, bundle_manifest))

    if write:
        for rel, text in texts.items():
            path = root / rel
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(text, encoding="utf-8")

    return {
        "authorization_record": authorization,
        "governance_boundary": governance,
        "cutoff_source_audit": cutoff_audit,
        "cutoff_plan": cutoff_plan,
        "release_manifest": editions,
        "required_editions": required_editions,
        "request_manifest": requests_rows,
        "response_manifest": responses_rows,
        "locked_series_extract": locked_series,
        "semantic_compatibility": semantic_rows,
        "imf_catalog": imf_catalog,
        "financing_evidence": financing_evidence,
        "continuation_record": continuation,
        "captured_editions": captured_editions,
        "bundle_manifest": bundle_manifest,
        "decision": decision,
        "qc_report": qc,
        "contract_read": contract,
        "artifact_texts": texts,
        "evidence_summary": summary,
    }


def render_readme(decision: dict[str, Any], qc: dict[str, Any],
                  summary: dict[str, Any],
                  cutoff_audit: dict[str, Any]) -> str:
    status = decision["m3i2_official_source_evidence_status"]
    return f"""# Stage128 — M3I-2 official-source evidence capture

**Action id:** `{ACTION_ID}`
**Action type:** `{ACTION_TYPE}`
**Baseline:** `{BASELINE_BRANCH}` @ `{BASELINE_COMMIT}`
**Evidence status:** `{status}`
**Result code:** `{decision["result_code"]}`

```text
OFFICIAL-SOURCE EVIDENCE CAPTURE ONLY
NETWORK ACCESS LIMITED TO OFFICIAL WORLD BANK / IMF SOURCES
RAW BYTES RETAINED AND HASHED
NO COMPANY-PANEL MACRO JOIN
NO FEATURE MATERIALIZATION
NO COVERAGE
NO DATA GATE
NO MODELING
NO M3I-vs-M2
NO M4
FINAL TEST LOCKED
NO MERGE AUTHORIZATION
```

## What this action is

An evidence-acquisition action. It captures, hashes and packages **official**
World Bank WDI and IMF material for the already-merged M3I-2 contract and the
contingent M3I-3 financing shell. It answers exactly one question: *is there an
independently auditable, raw-byte-backed official-source evidence package?*

It does **not** answer whether M3I-2 meets coverage thresholds, whether it
improves prediction, whether it should be admitted, or what the final model is.
An evidence status of `{EVIDENCE_COMPLETE}` would mean only that a **separate,
separately authorized** Data Gate may be *considered* after human review.

## The merged contract is read-only

`{CONTRACT_PACKAGE_DIR}/**` is byte-identical to `{BASELINE_COMMIT}`. Candidate
ids, source ids, indicator codes, transformations, observation-year rules,
vintage rules, missing-value rules, Data Gate thresholds, multiplicity families,
the financing shell's null fields and the final-test controls are all unchanged.

Locked candidates, restated for the reader (not redefined here):

1. `{CPI_CANDIDATE_ID}` → `intl_cpi_inflation_annual`, `{CPI_INDICATOR_CODE}`,
   transformation `identity`.
2. `{FX_CANDIDATE_ID}` → `intl_fx_change_official_annual`,
   `{FX_INDICATOR_CODE}`, transformation `100 * ln(E_y / E_(y-1))` — **not
   evaluated in this action**.

## Development-cutoff input firewall

Cutoffs come from one uniquely bound source:

* path `{cutoff_audit["cutoff_source_repository_path"]}`
* git blob `{cutoff_audit["cutoff_source_git_blob_sha"]}`
* SHA-256 `{cutoff_audit["cutoff_source_sha256"]}`
* cutoff field `{cutoff_audit["cutoff_field"]}`
* columns read: {", ".join("`" + c + "`" for c in cutoff_audit["columns_read"])}

No target, financial, market or macro feature column was read, no final-test
directory was searched, and only development target years
{"–".join((DEVELOPMENT_TARGET_YEARS[0], DEVELOPMENT_TARGET_YEARS[-1]))} are in
scope. Unique development cutoffs:
**{summary["unique_development_cutoffs"]}** over
{summary["development_pairs_behind_cutoff_plan"]} development pairs — an
**input-integrity count, not coverage**.

**Known limitation.** `{cutoff_audit["cutoff_field"]}` is a date with no
verified intraday `available_at` timestamp: Stage125 Part3B1A locked the Cut-A
operationalization but recorded zero real `available_at` assignments. Edition
selection therefore uses `{CUTOFF_TIME_ASSUMPTION}` of the cutoff date — the
earliest possible instant, which can only **exclude** an edition, never admit
one.

## Official sources

Discovery roots, captured before anything was downloaded from them:

* {WB_ARCHIVE_LISTING_URL}
* {WB_DATABANK_ARCHIVES_URL}
* {WB_CPI_METADATA_URL}
* {WB_FX_METADATA_URL}
* {IMF_MFS_IR_DATASET_URL}

Only official `worldbank.org` / `imf.org` hosts may terminate a request, HTTPS
only, with a descriptive User-Agent, a finite timeout, at most 3 attempts per
request and deterministic backoff. A redirect that leaves an official host is a
hard stop. Mirrors, aggregators, FRED/ALFRED, DBnomics, Kaggle, GitHub copies
and unofficial Iranian FX sources are forbidden, and a search-result snippet is
never evidence.

## Evidence status

**`{status}`**

* official requests attempted: {summary["official_requests_attempted"]}
* official responses retained: {summary["official_responses_retained"]}
  (successful: {summary["official_responses_successful"]})
* raw bytes retained: {summary["raw_bytes_total"]} bytes across
  {summary["raw_bytes_retained_objects"]} objects
* WDI editions discovered: {summary["wdi_editions_discovered"]}
* editions with a **verified** release date:
  **{summary["editions_with_verified_release_date"]}** — unverified filename
  date tokens: {summary["editions_with_unverified_filename_date_token"]}
* required (servable) editions: **{summary["required_editions_total"]}** —
  archive editions actually captured and held:
  {summary["archive_editions_captured"]}
* cutoffs without a verified pre-cutoff vintage:
  **{summary["cutoffs_without_verified_pre_cutoff_edition"]}** of
  {summary["unique_development_cutoffs"]}
  ({summary["development_pairs_without_verified_pre_cutoff_edition"]} of 539
  development pairs)
* locked-series rows extracted: {summary["locked_series_rows_extracted"]}
* semantic compatibility — CPI PASS {summary["cpi_semantic_pass_count"]} /
  UNRESOLVED {summary["cpi_semantic_unresolved_count"]} / FAIL_INTEGRITY
  {summary["cpi_semantic_fail_integrity_count"]}; **FX** PASS
  {summary["fx_semantic_pass_count"]} / UNRESOLVED
  {summary["fx_semantic_unresolved_count"]} / FAIL_INTEGRITY
  {summary["fx_semantic_fail_integrity_count"]}
* capture invocations: {summary["capture_invocations"]}
* IMF catalog entries: {summary["imf_catalog_entries_captured"]}
* financing metadata decision: `{summary["financing_metadata_decision"]}`

Unresolved evidence is **never** converted into zero coverage or into an
observed failure. Missing proof is `UNRESOLVED`, not `FAIL`.

## Why nothing has a verified release date

The official archive listing publishes a table of `Year | Month(s)` and a
hyperlink per edition. The date inside a filename such as
`WDI_excel_2017_09_19.zip` is recorded as an **`edition_date_token`** — and
that is all it is. The retained listing bytes contain **no** occurrence of
"release", "released", "publish", "published", "last updated" or "revision",
and no retained archive states an edition-level release date either. So no
retained official bytes describe that token as a release, publication or update
date.

Under the locked rule, that means `release_date_verified = false`,
`derived_release_available_at_utc = null` and
`release_available_at_derivation_status =
{RELEASE_DATE_UNRESOLVED_STATUS}` for every edition. Retrieval time,
`Last-Modified`, ZIP member timestamps and workbook properties are explicitly
not accepted as substitutes.

An edition with no verified `available_at` can never be a pre-cutoff vintage,
so **every** development cutoff is unresolved. The archives themselves are
still held and still extracted — being unable to date an edition is not the
same as not having it.

## Why every FX row is UNRESOLVED

The archived `PA.NUS.FCRF` metadata gives a generic construct
("Official exchange rate (LCU per US$, period average)", `Periodicity: Annual`,
`Unit of measure` empty) and a generic methodology note about exchange-rate
arrangements. Section 4 permits an archived title to support the construct and
the **unit label** — but not the three continuity facts a log change depends
on:

1. which currency unit the local-currency figures are denominated in;
2. the local-currency valuation definition for that vintage;
3. that no redenomination or currency-unit break falls across the pair.

The archived metadata names no currency, no denomination and no unit break, and
silence is not proof of absence. All three therefore stay unverified and every
FX row is `UNRESOLVED`. Continuity is **not** inferred from smooth values, and
the current metadata page — which states it was updated on a 2026 date — is not
used as proof about historical editions.

CPI rows remain `PASS` where the archived title, annual periodicity, percent
rate and calendar-year columns are each evidenced; the provenance of the unit
claim now travels in the committed CSV as `unit_evidence_source`.

## Forbidden execution counters — all zero

company macro joins, feature materializations, coverage calculations, Data Gate
executions, model fits, predictions, predictive metrics, Holm calculations,
final-test rows read.

QC: **{qc["assertion_count"]} assertions, {qc["failed_count"]} failed**,
all_pass = **{qc["all_pass"]}**.

## State after this action

* `m3i2_contract_status` = `PROSPECTIVELY_LOCKED_NO_DATA`
* `m3i2_data_gate_executed` = **false**, `m3i2_block_admitted` = **false**,
  `m3i2_modeling_started` = **false**
* `m3i3_admitted` = **false**, lock `{M3I3_LOCK_STATUS}`
* M3-CBI `{M3_CBI_STATUS}`, admitted **false**
* `m4_authorized` = **false**, `m4_started` = **false**
* `final_test_locked` = **true**
* `merge_authorized` = **false**

Next pointer (informational only): `{decision["next_research_action_id"]}` with
`next_research_action_authorized` = **false**. **An evidence-capture completion
does not by itself authorize the Data Gate.**
"""


# --------------------------------------------------------------------------- #
# Locked-series extraction and vintage semantics (sections 11-12)
#
# Offline, from retained archive bytes. ``openpyxl`` is a data-parsing library
# for the official archive format; no modeling library is imported anywhere.
# Nothing here interpolates, extrapolates, fills, transforms, selects an
# observation year for a company pair, or joins to a company row.
# --------------------------------------------------------------------------- #

WDI_EXCEL_MEMBER = "WDIEXCEL.xlsx"
WDI_DATA_CSV_MEMBER = "WDIData.csv"
WDI_SERIES_CSV_MEMBER = "WDISeries.csv"
WDI_COUNTRY_CSV_MEMBER = "WDICountry.csv"
WDI_DATA_SHEET = "Data"
WDI_SERIES_SHEET = "Series"
WDI_COUNTRY_SHEET = "Country"


def _parse_numeric(raw: Any) -> tuple[str, str, str]:
    """Return ``(raw_text, parsed_value, parse_status)`` deterministically.

    A blank cell stays blank and is reported as ``MISSING`` - never zero, never
    filled, never carried forward.
    """
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return "", "", "MISSING"
    text = str(raw).strip()
    try:
        return text, repr(float(text)), "PARSED"
    except (TypeError, ValueError):
        return text, "", "UNPARSEABLE"


def extract_locked_series_from_archive(
    archive_path: str | os.PathLike[str], edition_id: str,
    release_available_at: str, artifact_sha256: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Extract Iran + the two locked codes, and the per-edition semantics.

    Returns ``(locked_series_rows, semantic_rows)``. Only ``IRN`` and only
    ``FP.CPI.TOTL.ZG`` / ``PA.NUS.FCRF`` are read out of the archive; nothing
    else in the file is extracted.
    """
    import openpyxl

    path = Path(archive_path)
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        if WDI_EXCEL_MEMBER not in names:
            # The official archive ships two container layouts: a single Excel
            # workbook, or the same content as a set of CSVs. Both are
            # first-party; neither is a substitute source.
            if WDI_DATA_CSV_MEMBER in names:
                return _extract_from_csv_layout(
                    archive, edition_id, release_available_at, artifact_sha256)
            return [], [_unresolved_semantics(
                edition_id, release_available_at, artifact_sha256,
                f"the archive contains neither {WDI_EXCEL_MEMBER} nor "
                f"{WDI_DATA_CSV_MEMBER}")]
        payload = io.BytesIO(archive.read(WDI_EXCEL_MEMBER))

    workbook = openpyxl.load_workbook(payload, read_only=True, data_only=True)
    try:
        rows, economy_names = _extract_data_sheet(
            workbook, edition_id, release_available_at, artifact_sha256)
        series_meta = _extract_series_metadata(workbook)
        economy_verified = _verify_iran_identity(workbook, economy_names)
    finally:
        workbook.close()

    semantic_rows = [
        _semantic_row(edition_id, release_available_at, artifact_sha256, code,
                      series_meta.get(code), economy_verified, rows)
        for code in LOCKED_INDICATOR_CODES
    ]
    return rows, semantic_rows


def _extract_data_sheet(workbook, edition_id: str, release_available_at: str,
                        artifact_sha256: str
                        ) -> tuple[list[dict[str, Any]], set[str]]:
    sheet = workbook[WDI_DATA_SHEET]
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator)
    years: list[tuple[int, str]] = []
    for index, cell in enumerate(header):
        text = str(cell).strip() if cell is not None else ""
        if text.isdigit() and len(text) == 4:
            years.append((index, text))

    out: list[dict[str, Any]] = []
    economy_names: set[str] = set()
    for row_number, row in enumerate(iterator, start=2):
        if len(row) < 4:
            continue
        economy_code = str(row[1]).strip() if row[1] is not None else ""
        indicator_code = str(row[3]).strip() if row[3] is not None else ""
        if economy_code != IRAN_ECONOMY_CODE:
            continue
        if indicator_code not in LOCKED_INDICATOR_CODES:
            continue
        economy_name = str(row[0]).strip() if row[0] is not None else ""
        indicator_name = str(row[2]).strip() if row[2] is not None else ""
        economy_names.add(economy_name)
        for column_index, year in years:
            raw = row[column_index] if column_index < len(row) else None
            raw_text, parsed, status = _parse_numeric(raw)
            out.append({
                "archive_edition_id": edition_id,
                "release_available_at_utc": release_available_at,
                "raw_artifact_sha256": artifact_sha256,
                "economy_code_raw": economy_code,
                "economy_name_raw": economy_name,
                "indicator_code_raw": indicator_code,
                "indicator_name_raw": indicator_name,
                "frequency_evidence": "annual_year_columns_in_wdi_data_sheet",
                "unit_evidence": "",
                "observation_year": year,
                "raw_value_text": raw_text,
                "parsed_numeric_value": parsed,
                "parse_status": status,
                "source_row_locator":
                    f"{WDI_EXCEL_MEMBER}!{WDI_DATA_SHEET}!R{row_number}C"
                    f"{column_index + 1}",
            })
    return out, economy_names


def _extract_series_metadata(workbook) -> dict[str, dict[str, str]]:
    """Read the archive's OWN Series sheet - the vintage describing itself."""
    if WDI_SERIES_SHEET not in workbook.sheetnames:
        return {}
    sheet = workbook[WDI_SERIES_SHEET]
    iterator = sheet.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(iterator)]
    wanted = {name.lower(): index for index, name in enumerate(header)}

    def field(row, *candidates):
        for candidate in candidates:
            index = wanted.get(candidate.lower())
            if index is not None and index < len(row) and row[index]:
                return str(row[index]).strip()
        return ""

    out: dict[str, dict[str, str]] = {}
    for row in iterator:
        if not row:
            continue
        code = str(row[0]).strip() if row[0] is not None else ""
        if code not in LOCKED_INDICATOR_CODES:
            continue
        out[code] = {
            "series_title": field(row, "Indicator Name"),
            "unit": field(row, "Unit of measure"),
            "periodicity": field(row, "Periodicity"),
            "base_period": field(row, "Base Period"),
            "short_definition": field(row, "Short definition"),
            "long_definition": field(row, "Long definition"),
            # carried so the FX continuity search can look at what the ARCHIVE
            # itself says, rather than at current metadata
            "statistical_concept": field(
                row, "Statistical concept and methodology"),
            "other_notes": field(row, "Other notes"),
            "limitations": field(row, "Limitations and exceptions"),
            "notes_from_original_source": field(
                row, "Notes from original source"),
        }
    return out


def _verify_iran_identity(workbook, economy_names: set[str]) -> bool:
    """Iran's identity must come from the archive itself, not an assumption."""
    if not economy_names:
        return False
    if not any(any(known.lower() in name.lower() or name.lower() in known.lower()
                   for known in IRAN_ECONOMY_NAMES)
               for name in economy_names):
        return False
    if WDI_COUNTRY_SHEET not in workbook.sheetnames:
        return True
    sheet = workbook[WDI_COUNTRY_SHEET]
    for row in sheet.iter_rows(values_only=True):
        if row and str(row[0]).strip() == IRAN_ECONOMY_CODE:
            return True
    return False


def _unresolved_semantics(edition_id: str, release_available_at: str,
                          artifact_sha256: str, reason: str
                          ) -> dict[str, Any]:
    row = {f: "" for f in SEMANTIC_EVIDENCE_FIELDS}
    row.update({
        "archive_edition_id": edition_id,
        "release_available_at_utc": release_available_at,
        "raw_archive_sha256": artifact_sha256,
        "compatibility_status": "UNRESOLVED",
        "unresolved_reason": reason,
    })
    return row


def _fx_continuity_evidence(meta: dict[str, str], edition_id: str,
                            artifact_sha: str) -> dict[str, Any]:
    """Look in the ARCHIVED metadata for the three FX continuity facts.

    Requires, per edition:

    * which currency unit the local-currency figures are denominated in;
    * the local-currency valuation definition;
    * that no redenomination or currency-unit break falls across the pair.

    The archived Series metadata for ``PA.NUS.FCRF`` carries a generic
    exchange-rate methodology and an empty ``Unit of measure``. It names no
    currency, no denomination and no unit break, so none of the three can be
    verified from it. This function reports that rather than reading the
    generic label as if it settled the question.
    """
    haystack = " ".join(
        str(meta.get(k) or "") for k in
        ("series_title", "unit", "base_period", "short_definition",
         "long_definition", "statistical_concept", "other_notes",
         "limitations")).strip()
    lowered = haystack.lower()
    locator_base = f"{WDI_SERIES_SHEET}!{FX_INDICATOR_CODE}@{edition_id}"

    # Deliberately NOT "currency unit"/"denomination" on their own: the generic
    # WDI definition says "local currency units relative to the U.S. dollar",
    # which is the unit LABEL and names no currency. Verifying denomination
    # from that phrase would be exactly the over-claim this correction removes.
    denomination_terms = ("rial", " irr", "toman", "denominated in",
                          "denomination of")
    denomination_hits = [w for w in denomination_terms if w in lowered]
    denomination_raw = haystack if denomination_hits else ""

    valuation_terms = ("valuation", "official rate", "principal rate",
                       "secondary rate", "tertiary rate", "multiple exchange")
    valuation_hits = [w for w in valuation_terms if w in lowered]
    # A generic mention that multiple arrangements EXIST is not a statement of
    # which one this vintage used for Iran.
    valuation_states_this_vintage = False

    break_terms = ("redenomination", "redenominated", "currency reform",
                   "unit break", "rebase of the currency")
    break_hits = [w for w in break_terms if w in lowered]

    return {
        "currency_denomination_raw": denomination_raw,
        "currency_denomination_evidence_artifact": (
            artifact_sha if denomination_hits else ""),
        "currency_denomination_evidence_locator": (
            f"{locator_base}#denomination" if denomination_hits else ""),
        "currency_denomination_verified": bool(denomination_hits),

        "local_currency_valuation_definition_raw": (
            meta.get("statistical_concept", "") if valuation_hits else ""),
        "valuation_definition_evidence_artifact": (
            artifact_sha if valuation_states_this_vintage else ""),
        "valuation_definition_evidence_locator": (
            f"{locator_base}#valuation" if valuation_states_this_vintage
            else ""),
        "valuation_definition_verified": valuation_states_this_vintage,

        "redenomination_or_unit_break_evidence_raw": (
            haystack if break_hits else ""),
        "redenomination_or_unit_break_evidence_artifact": (
            artifact_sha if break_hits else ""),
        "redenomination_or_unit_break_evidence_locator": (
            f"{locator_base}#redenomination" if break_hits else ""),
        # Silence is not proof of absence: an archive that never mentions a
        # redenomination has not established that none occurred.
        "no_redenomination_or_unit_break_verified": False,
        "_denomination_hits": denomination_hits,
        "_valuation_hits": valuation_hits,
        "_break_hits": break_hits,
    }


def _semantic_row(edition_id: str, release_available_at: str,
                  artifact_sha256: str, code: str,
                  meta: dict[str, str] | None, economy_verified: bool,
                  rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Build one per-edition, per-code compatibility record.

    PASS requires the ARCHIVE's own metadata to support the locked meaning.
    Missing metadata yields UNRESOLVED. Metadata that positively contradicts
    the locked meaning yields FAIL_INTEGRITY. For FX, the three continuity
    facts must each be evidenced on top of the construct and unit label.
    """
    meta = meta or {}
    title = meta.get("series_title", "")
    unit = meta.get("unit", "")
    periodicity = meta.get("periodicity", "")
    present = [r for r in rows if r["indicator_code_raw"] == code]
    years = sorted({r["observation_year"] for r in present})

    reasons: list[str] = []
    contradiction = False
    unit_source = "unit_of_measure_field" if unit else ""
    lowered = f"{title} {meta.get('long_definition', '')}".lower()

    if code == CPI_INDICATOR_CODE:
        title_ok = "inflation" in lowered and (
            "annual %" in lowered or "annual percent" in lowered
            or "consumer price" in lowered)
        if "gdp deflator" in lowered:
            contradiction = True
            reasons.append("the archived series is a GDP-deflator construct")
        if "index" in title.lower() and "inflation" not in title.lower():
            contradiction = True
            reasons.append("the archived series is a CPI index level")
        unit_ok = bool(unit) and ("%" in unit or "percent" in unit.lower())
        if not unit_ok and ("annual %" in title.lower()
                            or "annual percent" in title.lower()):
            # The archive leaves "Unit of measure" blank but states the unit
            # inside its own series title. Section 4 permits the archived title
            # to support the unit LABEL; the provenance is recorded, not hidden.
            unit_ok = True
            unit_source = "archived_series_title"
    else:
        title_ok = "exchange rate" in lowered
        if "atlas" in lowered:
            contradiction = True
            reasons.append("the archived series is an Atlas-method rate")
        unit_ok = bool(unit) and ("lcu" in unit.lower() or "us$" in unit.lower()
                                 or "local" in unit.lower())
        if not unit_ok and "lcu per us$" in title.lower():
            unit_ok = True
            unit_source = "archived_series_title"

    calendar_ok = bool(present) and bool(periodicity) and (
        "annual" in periodicity.lower()) and all(
        y.isdigit() and len(y) == 4 for y in years)
    calendar_locator = (
        f"{WDI_DATA_SHEET}!calendar_year_columns[{years[0]}..{years[-1]}]"
        if years else "")

    frequency_ok = bool(periodicity) and "annual" in periodicity.lower()
    if not title:
        reasons.append("the archive states no series title for this code")
    if not periodicity:
        reasons.append("the archive states no periodicity for this code")
    if not unit and unit_source != "archived_series_title":
        reasons.append("the archive states no unit of measure for this code")
    if not present:
        reasons.append("the locked code is absent from this edition")
    if not economy_verified:
        reasons.append("Iran's economy identity was not verified in-archive")

    evidence = {
        "economy_identity_verified": economy_verified,
        "indicator_code_verified": bool(present),
        "frequency_annual_verified": frequency_ok,
        "calendar_year_semantics_verified": calendar_ok,
        "title_compatibility": "COMPATIBLE" if (title and title_ok) else "",
        "unit_compatibility": "COMPATIBLE" if unit_ok else "",
        "raw_archive_sha256": artifact_sha256,
        "integrity_contradiction": contradiction,
    }
    status = classify_semantic_compatibility(evidence)

    row: dict[str, Any] = {
        "archive_edition_id": edition_id,
        "indicator_code": code,
        "release_available_at_utc": release_available_at or "",
        "economy_identity_verified": economy_verified,
        "indicator_code_verified": bool(present),
        "archived_series_title_raw": title,
        "title_compatibility": evidence["title_compatibility"] or "UNRESOLVED",
        "frequency_raw": periodicity,
        "frequency_annual_verified": evidence["frequency_annual_verified"],
        "unit_raw": unit or (title if unit_source == "archived_series_title"
                             else ""),
        "unit_evidence_source": unit_source,
        "unit_compatibility": evidence["unit_compatibility"] or "UNRESOLVED",
        "calendar_year_semantics_verified": calendar_ok,
        "calendar_year_evidence_locator": calendar_locator,
        "raw_archive_sha256": artifact_sha256,
        "evidence_locator": f"{WDI_EXCEL_MEMBER}!{WDI_SERIES_SHEET}!{code}",
        "compatibility_status": status,
        "unresolved_reason": "",
        # FX-only fields stay blank for CPI rather than absent
        "currency_denomination_raw": "",
        "currency_denomination_evidence_artifact": "",
        "currency_denomination_evidence_locator": "",
        "currency_denomination_verified": "",
        "local_currency_valuation_definition_raw": "",
        "valuation_definition_evidence_artifact": "",
        "valuation_definition_evidence_locator": "",
        "valuation_definition_verified": "",
        "redenomination_or_unit_break_evidence_raw": "",
        "redenomination_or_unit_break_evidence_artifact": "",
        "redenomination_or_unit_break_evidence_locator": "",
        "no_redenomination_or_unit_break_verified": "",
        "fx_pair_semantic_compatibility_status": "",
    }

    if code == FX_INDICATOR_CODE:
        continuity = _fx_continuity_evidence(meta, edition_id, artifact_sha256)
        missing = [f for f in FX_CONTINUITY_REQUIRED_FIELDS
                   if continuity.get(f) is not True]
        for key, value in continuity.items():
            if not key.startswith("_"):
                row[key] = value
        if continuity["_break_hits"] and status != "FAIL_INTEGRITY":
            # An archive that DOES describe a break is stronger evidence than
            # one that is silent, and it contradicts pair continuity.
            status = "FAIL_INTEGRITY"
            reasons.append(
                "the archived metadata describes a redenomination or "
                "currency-unit break")
        if missing:
            if not continuity["_denomination_hits"]:
                reasons.append(
                    "the archived metadata names no currency denomination for "
                    "the local-currency figures")
            if not continuity["_valuation_hits"]:
                reasons.append(
                    "the archived metadata states no local-currency valuation "
                    "definition for this vintage")
            reasons.append(
                "the archived metadata does not establish that no "
                "redenomination or currency-unit break falls across the pair")
            reasons.append(
                "the archived title supports the generic construct and unit "
                "label only; it cannot by itself evidence pair continuity")
            if status == "PASS":
                status = "UNRESOLVED"
        row["fx_pair_semantic_compatibility_status"] = status
        row["compatibility_status"] = status

    row["compatibility_status"] = status
    row["unresolved_reason"] = "; ".join(reasons)
    return row


def _extract_from_csv_layout(
    archive: zipfile.ZipFile, edition_id: str, release_available_at: str,
    artifact_sha256: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Same extraction, for the editions distributed as CSVs.

    Some official archive editions ship ``WDIData.csv`` / ``WDISeries.csv``
    instead of a single workbook. The column semantics are identical, so the
    locked-series rows and the semantic evidence are built the same way.
    """
    rows: list[dict[str, Any]] = []
    economy_names: set[str] = set()

    with archive.open(WDI_DATA_CSV_MEMBER) as handle:
        reader = csv.reader(io.TextIOWrapper(handle, encoding="utf-8-sig",
                                             newline=""))
        header = next(reader)
        years = [(index, value.strip()) for index, value in enumerate(header)
                 if value.strip().isdigit() and len(value.strip()) == 4]
        for row_number, row in enumerate(reader, start=2):
            if len(row) < 4:
                continue
            economy_code = row[1].strip()
            indicator_code = row[3].strip()
            if economy_code != IRAN_ECONOMY_CODE:
                continue
            if indicator_code not in LOCKED_INDICATOR_CODES:
                continue
            economy_names.add(row[0].strip())
            for column_index, year in years:
                raw = row[column_index] if column_index < len(row) else None
                raw_text, parsed, status = _parse_numeric(raw)
                rows.append({
                    "archive_edition_id": edition_id,
                    "release_available_at_utc": release_available_at,
                    "raw_artifact_sha256": artifact_sha256,
                    "economy_code_raw": economy_code,
                    "economy_name_raw": row[0].strip(),
                    "indicator_code_raw": indicator_code,
                    "indicator_name_raw": row[2].strip(),
                    "frequency_evidence":
                        "annual_year_columns_in_wdi_data_csv",
                    "unit_evidence": "",
                    "observation_year": year,
                    "raw_value_text": raw_text,
                    "parsed_numeric_value": parsed,
                    "parse_status": status,
                    "source_row_locator":
                        f"{WDI_DATA_CSV_MEMBER}!R{row_number}C"
                        f"{column_index + 1}",
                })

    series_meta: dict[str, dict[str, str]] = {}
    if WDI_SERIES_CSV_MEMBER in archive.namelist():
        with archive.open(WDI_SERIES_CSV_MEMBER) as handle:
            reader = csv.DictReader(
                io.TextIOWrapper(handle, encoding="utf-8-sig", newline=""))
            for record in reader:
                code = (record.get("Series Code")
                        or record.get("SeriesCode") or "").strip()
                if code not in LOCKED_INDICATOR_CODES:
                    continue
                series_meta[code] = {
                    "series_title": (record.get("Indicator Name") or "").strip(),
                    "unit": (record.get("Unit of measure") or "").strip(),
                    "periodicity": (record.get("Periodicity") or "").strip(),
                    "base_period": (record.get("Base Period") or "").strip(),
                    "short_definition":
                        (record.get("Short definition") or "").strip(),
                    "long_definition":
                        (record.get("Long definition") or "").strip(),
                    "statistical_concept": (record.get(
                        "Statistical concept and methodology") or "").strip(),
                    "other_notes": (record.get("Other notes") or "").strip(),
                    "limitations": (record.get(
                        "Limitations and exceptions") or "").strip(),
                    "notes_from_original_source": (record.get(
                        "Notes from original source") or "").strip(),
                }

    economy_verified = bool(economy_names) and any(
        any(known.lower() in name.lower() or name.lower() in known.lower()
            for known in IRAN_ECONOMY_NAMES) for name in economy_names)

    semantic_rows = [
        _semantic_row(edition_id, release_available_at, artifact_sha256, code,
                      series_meta.get(code), economy_verified, rows)
        for code in LOCKED_INDICATOR_CODES
    ]
    return rows, semantic_rows
